import pandas as pd
import numpy as np
import os
import glob
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from shutil import copyfile
import datetime
import warnings
from difflib import SequenceMatcher , get_close_matches

warnings.filterwarnings('ignore')






def gstchecksum(gst_no):


"""
This function gstchecksum will check the last digit of the gst number and return whether the Check sum matches or not

Output will be given as "Check Sum MATCH" or "Check Sum MISMATCH"

:param gst_no: This is the only argument that needs to be given. It is a mandatory Argument.The argument must be 15 digit long

:type gst_no: This parameter must be a string and must be 15 digit long

:return : The function will return only one of two values a) Check Sum MATCH & b) Check Sum MISMATCH

:raises: There are two errors that will be raised by the Function 
        1. Type Error: If the parameter entered is not a string, then this error is raised
        2. Exception: If the parameter entered is not 15 digit long, then Exception is raised

:see also: To know how the GST Ceck sum is calculated , see the alogorithm behind the last digit



"""

    charlist=[char for char in gst_no.upper()]

    a=1
    cumhash=[]

    if not type(gst_no) is str:
        raise TypeError("Only strings are allowed")


    try:
        if len(gst_no)<15:
            raise Exception("Please ensure that the input is 15 digit long")
    except:
        pass


    for i in charlist[0:14:1]:
        
        if a % 2==0:
            multiplier=2
        else:
            multiplier=1

        if i.isdigit():
            intvalue=int(i)
            prod=intvalue*multiplier
            quotient=prod//36
            remain=prod%36
            hash=quotient+remain
            
        else:
            intvalue=ord(i)-55   
            prod=intvalue*multiplier
            quotient=prod//36
            remain=prod%36
            hash=quotient+remain
            

        a=a+1

        cumhash.append(hash)

    hashsum=(sum(cumhash))

    remain=hashsum%36

    checksum=36-remain

    if checksum<10:
        finalchk=int(checksum)
    else:
        finalchk=chr(checksum+55)
   
    
    lastchr=(gst_no[len(gst_no)-1])
    
    if lastchr.isdigit():
        lastchr=int(lastchr)
    else:
        lastchr=str(lastchr)
        
        
    if finalchk==lastchr:
        result="Check Sum MATCH"
    else:
        result="Check Sum MISMATCH"
    
    return (result)
     


def gstinvcheck(a):

"""
This function will check whether the invoice number entered is correct or not.

As per GST rules, the Invoice number must be maximum 15 digit long

:param a: this must be the GST  Invoice number 

:param a type: The Type of parameter must be a string. However, in the functionit is converting any parameter into a string through str() method

:return : it return one of 2 output  a) Invoice Number Valid or b) Invoice Number Invalid



"""

    try:
        length=len(str(a))
    except:
        length=0    


    if length<=16:
        status="Invoice Number Valid"
    else:
        status="Invoice Number Invalid"
        
    return(status)





def extract_pan(gst_no):



"""
This fucntion will extract the PAN number from the provided GST_No

:param gst_no: This function requires only one parameter. ie the GST No

:param gst_no type: The type of the parameter must be a string

:return :The function will return a string which is the PAN Number

:SeeAlso : The PAN number is the 3rd Character to 12th Character of the GST Number



"""
    if not type(gst_no) is str:
        raise TypeError("Only strings are allowed")


    try:
        if len(gst_no)<15:
            raise Exception("Please ensure that the input is 15 digit long")
    except:
        pass

    

    try:
        pan_num=gst_no[2:12:1]
    except:
        pan_num=gst_no
    
    return(pan_num)



def gstr2a_merge(filepath):


"""
This is a super Useful Function for merging all the GSTR2A files kept in a folder.


:Param filepath: The function takes only one parameter which is the Compelte File Path to any one GSTR2A file in that folder

:Param type : The type of the argument should be the complete path to the excel file of the GSTR2A , till the extension 

:return : The function will return a merged Excel File by merging all the GSTR2A Sheets i.e the B2B, B2BA, CDNR, CDNRA

:Warnings:  Please note that only one argument needs to be passed
            
            The Function will auto read all other files which are there in the folder

            Please ensure that the particular folder only has the GSTR2A files that you want to combine. 

            There should not be any other files as this function will read all the files in that folder whether or not they are GSTR2A files

:See Also: Just a snippet of how this function works

            The code will forst loop through all the B2B files of the folder and then B2B A, then CDNR and then CDNRA.

            After looping through all the files, this will first make 4 different sheets of these 4 types

            Then it will make a All Combined Files by merging all the B2B, B2BA, CDNR and CDNRA

            Also, as an additional analysis, this will make 3 different sheets a) RCM Cases b) GSTR-1 Filing Status as No c) tax Zero cases

            NOw, as per the GST Act, and Rule 36, for claiming the ITC, GSTR -1 Fling status to be Y and RCM should be N



"""

    import pandas as pd
    import glob
    import os


    pth = os.path.dirname(filepath)

    filenames = glob.glob(pth + "/*.xlsx")

    warnings.filterwarnings('ignore')

    

    cum_size = 0

    for file in filenames:
        size = os.path.getsize(file)

        cum_size = cum_size + size

        if size > 31457280:
            print("Please upload a smaller file size. Maximum limit is 30 mb.")

        elif cum_size > 314572800:
            print("Combined File size for all the file is more than 300 mb. Please use smaller files")
            break
        else:
            pass

    # A. iterate through each file to append it one below the other

    # A.1 : This will iterate through the B2B file


    print(f"The files that will be combined are {filenames}")


    print("We are working on B2B sheet of all the monthly GSTR2A..Please wait...")

    
    df2 = pd.DataFrame()

    for files in filenames:
        df = pd.read_excel(files, sheet_name=1)

        df1 = df.drop([0, 1, 2, 3, 4], axis=0)

        df1 = df1.dropna(how='all')

        df1['File_name'] = files

        df2 = df2.append(df1)

    # this is used for deleting all the rows which are totally blank

    df3 = df2

    # this is used for renaming the names of the columns

    df3.rename(columns={'Goods and Services Tax  - GSTR 2A': 'GSTIN_of_Supplier'}, inplace=True)
    df3.rename(columns={'Unnamed: 1': 'Legal_Name_Of Supplier'}, inplace=True)
    df3.rename(columns={'Unnamed: 2': 'Inv_CN_DN_Number_Original'}, inplace=True)
    df3.rename(columns={'Unnamed: 3': 'Inv_CN_DN_Type_Original'}, inplace=True)
    df3.rename(columns={'Unnamed: 4': 'Inv_CN_DN_Date_Original'}, inplace=True)
    df3.rename(columns={'Unnamed: 5': 'Inv_CN_DN_Value_Original'}, inplace=True)
    df3.rename(columns={'Unnamed: 6': 'Place_Of_Supply'}, inplace=True)
    df3.rename(columns={'Unnamed: 7': 'Supply_Attract_Reverse_Charge'}, inplace=True)
    df3.rename(columns={'Unnamed: 8': 'GST_Rate'}, inplace=True)
    df3.rename(columns={'Unnamed: 9': 'Taxable_Value_Rs'}, inplace=True)
    df3.rename(columns={'Unnamed: 10': 'IGST_Rs'}, inplace=True)
    df3.rename(columns={'Unnamed: 11': 'CGST_Rs'}, inplace=True)
    df3.rename(columns={'Unnamed: 12': 'SGST_Rs'}, inplace=True)
    df3.rename(columns={'Unnamed: 13': 'Cess'}, inplace=True)
    df3.rename(columns={'Unnamed: 14': 'GSTR_1_5_Filing_Status'}, inplace=True)
    df3.rename(columns={'Unnamed: 15': 'GSTR_1_5_Filing_Date'}, inplace=True)
    df3.rename(columns={'Unnamed: 16': 'GSTR_1_5_Filing_Period'}, inplace=True)
    df3.rename(columns={'Unnamed: 17': 'GSTR_3B_Filing_Status'}, inplace=True)
    df3.rename(columns={'Unnamed: 18': 'Amendment_made_if_any'}, inplace=True)
    df3.rename(columns={'Unnamed: 19': 'Tax_Period_in_which_Amended'}, inplace=True)
    df3.rename(columns={'Unnamed: 20': 'Effective_date_of_cancellation'}, inplace=True)
    df3.rename(columns={'Unnamed: 21': 'Source'}, inplace=True)
    df3.rename(columns={'Unnamed: 22': 'IRN'}, inplace=True)
    df3.rename(columns={'Unnamed: 23': 'IRN_Date'}, inplace=True)

    # here we will remove the rows, in which the invoice number has  a total
    filt = df3['Inv_CN_DN_Number_Original'].str.contains('Total', na=False)
    df3 = df3[~filt]

    df3['Inv_CN_DN_Date_Text'] = df3['Inv_CN_DN_Date_Original'].str.replace("-", ".")
    df3['Total_Tax'] = df3['IGST_Rs'] + df3['CGST_Rs'] + df3['SGST_Rs']
    df3['Unique_ID'] = df3['GSTIN_of_Supplier'] + "/" + df3['Inv_CN_DN_Number_Original'] + "/" + df3[
        'Inv_CN_DN_Date_Text']

    df3['Sheet_Name'] = ("B2B")

    df3['PAN_Number'] = df3["GSTIN_of_Supplier"].apply(lambda x: x[2:12:1])

    df3 = df3.replace(np.nan, "", regex=True)


    # A.2 : This will iterate through the B2BA file

    print("We are working on B2BA sheet of all the monthly GSTR2A..Please wait")

    df2 = pd.DataFrame()

    for files in filenames:
        df = pd.read_excel(files, sheet_name=2)

        df1 = df.drop([0, 1, 2, 3, 4, 5], axis=0)

        df1 = df1.dropna(how='all')

        df1['File_name'] = files

        df2 = df2.append(df1)

    # this is used for deleting all the rows which are totally blank
    df4 = df2.dropna(how='all')

    # this is used for renaming the names of the columns

    df4.rename(
        columns={'                                      Goods and Services Tax - GSTR-2A': 'Inv_CN_DN_Number_Original'},
        inplace=True)
    df4.rename(columns={'Unnamed: 1': 'Inv_CN_DN_Date_Original'}, inplace=True)
    df4.rename(columns={'Unnamed: 2': 'GSTIN_of_Supplier'}, inplace=True)
    df4.rename(columns={'Unnamed: 3': 'Legal_Name_Of Supplier'}, inplace=True)
    df4.rename(columns={'Unnamed: 4': 'Inv_CN_DN_Type_Revised'}, inplace=True)
    df4.rename(columns={'Unnamed: 5': 'Inv_CN_DN_Number_Revised'}, inplace=True)
    df4.rename(columns={'Unnamed: 6': 'Inv_CN_DN_Date_Revised'}, inplace=True)
    df4.rename(columns={'Unnamed: 7': 'Inv_CN_DN_Value_Revised'}, inplace=True)
    df4.rename(columns={'Unnamed: 8': 'Place_Of_Supply'}, inplace=True)
    df4.rename(columns={'Unnamed: 9': 'Supply_Attract_Reverse_Charge'}, inplace=True)
    df4.rename(columns={'Unnamed: 10': 'GST_Rate'}, inplace=True)
    df4.rename(columns={'Unnamed: 11': 'Taxable_Value_Rs'}, inplace=True)
    df4.rename(columns={'Unnamed: 12': 'IGST_Rs'}, inplace=True)
    df4.rename(columns={'Unnamed: 13': 'CGST_Rs'}, inplace=True)
    df4.rename(columns={'Unnamed: 14': 'SGST_Rs'}, inplace=True)
    df4.rename(columns={'Unnamed: 15': 'Cess'}, inplace=True)
    df4.rename(columns={'Unnamed: 16': 'GSTR_1_5_Filing_Status'}, inplace=True)
    df4.rename(columns={'Unnamed: 17': 'GSTR_1_5_Filing_Date'}, inplace=True)
    df4.rename(columns={'Unnamed: 18': 'GSTR_1_5_Filing_Period'}, inplace=True)
    df4.rename(columns={'Unnamed: 19': 'GSTR_3B_Filing_Status'}, inplace=True)
    df4.rename(columns={'Unnamed: 20': 'Effective_date_of_cancellation'}, inplace=True)
    df4.rename(columns={'Unnamed: 21': 'Amendment_made_if_any'}, inplace=True)
    df4.rename(columns={'Unnamed: 22': 'Original_tax_period_in_which_reported'}, inplace=True)

    # here we will remove the rows, in which the invoice number has  a total
    filt = df4['Inv_CN_DN_Number_Revised'].str.contains('Total', na=False)

    df4 = df4[~filt]

    df4['Inv_CN_DN_Date_Text'] = df4['Inv_CN_DN_Date_Original'].str.replace("-", ".")
    df4['Total_Tax'] = df4['IGST_Rs'] + df4['CGST_Rs'] + df4['SGST_Rs']
    df4['Unique_ID'] = df4['GSTIN_of_Supplier'] + "/" + df4['Inv_CN_DN_Number_Original'] + "/" + df4[
        'Inv_CN_DN_Date_Text']
    df4["Inv_CN_DN_Date_Revised_Unique"] = df4['Inv_CN_DN_Date_Revised'].str.replace("-", ".")

    df4['Sheet_Name'] = ("B2BA")

    df4['PAN_Number'] = df4["GSTIN_of_Supplier"].apply(lambda x: x[2:12:1])

    df4 = df4.replace(np.nan, "", regex=True)

    # A.3 : This will iterate through the CDNR file

    print("We are working on CDNR sheet of all the monthly GSTR2A..Please wait")

    df2 = pd.DataFrame()

    for files in filenames:
        df = pd.read_excel(files, sheet_name=3)

        df1 = df.drop([0, 1, 2, 3, 4], axis=0)

        df1 = df1.dropna(how='all')

        df1['File_name'] = files

        df2 = df2.append(df1)

    # this is used for deleting all the rows which are totally blank
    df5 = df2.dropna(how='all')

    # this is used for renaming the names of the columns

    df5.rename(
        columns={'                                             Goods and Services Tax - GSTR-2A': 'GSTIN_of_Supplier'},
        inplace=True)
    df5.rename(columns={'Unnamed: 1': 'Legal_Name_Of Supplier'}, inplace=True)
    df5.rename(columns={'Unnamed: 2': 'Credit_Debit_Note_Original'}, inplace=True)
    df5.rename(columns={'Unnamed: 3': 'Inv_CN_DN_Number_Original'}, inplace=True)
    df5.rename(columns={'Unnamed: 4': 'Inv_CN_DN_Type_Original'}, inplace=True)
    df5.rename(columns={'Unnamed: 5': 'Inv_CN_DN_Date_Original'}, inplace=True)
    df5.rename(columns={'Unnamed: 6': 'Inv_CN_DN_Value_Original'}, inplace=True)
    df5.rename(columns={'Unnamed: 7': 'Place_Of_Supply'}, inplace=True)
    df5.rename(columns={'Unnamed: 8': 'Supply_Attract_Reverse_Charge'}, inplace=True)
    df5.rename(columns={'Unnamed: 9': 'GST_Rate'}, inplace=True)
    df5.rename(columns={'Unnamed: 10': 'Taxable_Value_Rs'}, inplace=True)
    df5.rename(columns={'Unnamed: 11': 'IGST_Rs'}, inplace=True)
    df5.rename(columns={'Unnamed: 12': 'CGST_Rs'}, inplace=True)
    df5.rename(columns={'Unnamed: 13': 'SGST_Rs'}, inplace=True)
    df5.rename(columns={'Unnamed: 14': 'Cess'}, inplace=True)
    df5.rename(columns={'Unnamed: 15': 'GSTR_1_5_Filing_Status'}, inplace=True)
    df5.rename(columns={'Unnamed: 16': 'GSTR_1_5_Filing_Date'}, inplace=True)
    df5.rename(columns={'Unnamed: 17': 'GSTR_1_5_Filing_Period'}, inplace=True)
    df5.rename(columns={'Unnamed: 18': 'GSTR_3B_Filing_Status'}, inplace=True)
    df5.rename(columns={'Unnamed: 19': 'Amendment_made_if_any'}, inplace=True)
    df5.rename(columns={'Unnamed: 20': 'Tax_Period_in_which_Amended'}, inplace=True)
    df5.rename(columns={'Unnamed: 21': 'Effective_date_of_cancellation'}, inplace=True)
    df5.rename(columns={'Unnamed: 22': 'Source'}, inplace=True)
    df5.rename(columns={'Unnamed: 23': 'IRN'}, inplace=True)
    df5.rename(columns={'Unnamed: 24': 'IRN_Date'}, inplace=True)

    # here we will remove the rows, in which the invoice number has  a total
    filt = df5['Inv_CN_DN_Number_Original'].str.contains('Total', na=False)

    df5 = df5[~filt]

    df5['Inv_CN_DN_Date_Text'] = df5['Inv_CN_DN_Date_Original'].str.replace("-", ".")
    df5['Total_Tax'] = df5['IGST_Rs'] + df5['CGST_Rs'] + df5['SGST_Rs']
    df5['Unique_ID'] = df5['GSTIN_of_Supplier'] + "/" + df5['Inv_CN_DN_Number_Original'] + "/" + df5[
        'Inv_CN_DN_Date_Text']

    df5['Sheet_Name'] = ("CDNR")

    df5['PAN_Number'] = df5["GSTIN_of_Supplier"].apply(lambda x: x[2:12:1])

    df5 = df5.replace(np.nan, "", regex=True)

    # A.2 : This will iterate through the CDNRA file

    print("We are working on CDNRA sheet of all the monthly GSTR2A..Please wait")

    df2 = pd.DataFrame()

    for files in filenames:
        df = pd.read_excel(files, sheet_name=4)

        df1 = df.drop([0, 1, 2, 3, 4, 5], axis=0)

        df1 = df1.dropna(how='all')

        df1['File_name'] = files

        df2 = df2.append(df1)

    # this is used for deleting all the rows which are totally blank
    df6 = df2.dropna(how='all')

    # this is used for renaming the names of the columns

    df6.rename(columns={'                             Goods and Services Tax - GSTR2A': 'Credit_Debit_Note_Original'},
               inplace=True)
    df6.rename(columns={'Unnamed: 1': 'Inv_CN_DN_Number_Original'}, inplace=True)
    df6.rename(columns={'Unnamed: 2': 'Inv_CN_DN_Date_Original'}, inplace=True)
    df6.rename(columns={'Unnamed: 3': 'GSTIN_of_Supplier'}, inplace=True)
    df6.rename(columns={'Unnamed: 4': 'Legal_Name_Of Supplier'}, inplace=True)
    df6.rename(columns={'Unnamed: 5': 'Credit_Debit_Note_Revised'}, inplace=True)
    df6.rename(columns={'Unnamed: 6': 'Inv_CN_DN_Number_Revised'}, inplace=True)
    df6.rename(columns={'Unnamed: 7': 'Inv_CN_DN_Type_Revised'}, inplace=True)
    df6.rename(columns={'Unnamed: 8': 'Inv_CN_DN_Date_Revised'}, inplace=True)
    df6.rename(columns={'Unnamed: 9': 'Inv_CN_DN_Value_Revised'}, inplace=True)
    df6.rename(columns={'Unnamed: 10': 'Place_Of_Supply'}, inplace=True)
    df6.rename(columns={'Unnamed: 11': 'Supply_Attract_Reverse_Charge'}, inplace=True)
    df6.rename(columns={'Unnamed: 12': 'GST_Rate'}, inplace=True)
    df6.rename(columns={'Unnamed: 13': 'Taxable_Value_Rs'}, inplace=True)
    df6.rename(columns={'Unnamed: 14': 'IGST_Rs'}, inplace=True)
    df6.rename(columns={'Unnamed: 15': 'CGST_Rs'}, inplace=True)
    df6.rename(columns={'Unnamed: 16': 'SGST_Rs'}, inplace=True)
    df6.rename(columns={'Unnamed: 17': 'Cess'}, inplace=True)
    df6.rename(columns={'Unnamed: 18': 'GSTR_1_5_Filing_Status'}, inplace=True)
    df6.rename(columns={'Unnamed: 19': 'GSTR_1_5_Filing_Date'}, inplace=True)
    df6.rename(columns={'Unnamed: 20': 'GSTR_1_5_Filing_Period'}, inplace=True)
    df6.rename(columns={'Unnamed: 21': 'GSTR_3B_Filing_Status'}, inplace=True)
    df6.rename(columns={'Unnamed: 22': 'Amendment_made_if_any'}, inplace=True)
    df6.rename(columns={'Unnamed: 23': 'Original_tax_period_in_which_reported'}, inplace=True)
    df6.rename(columns={'Unnamed: 24': 'Effective_date_of_cancellation'}, inplace=True)

    # here we will remove the rows, in which the invoice number has  a total
    filt = df6['Inv_CN_DN_Number_Revised'].str.contains('Total', na=False)

    df6 = df6[~filt]

    df6['Inv_CN_DN_Date_Text'] = df6['Inv_CN_DN_Date_Original'].str.replace("-", ".")
    df6['Total_Tax'] = df6['IGST_Rs'] + df6['CGST_Rs'] + df6['SGST_Rs']
    df6['Unique_ID'] = df6['GSTIN_of_Supplier'] + "/" + df6['Inv_CN_DN_Number_Original'] + "/" + df6[
        'Inv_CN_DN_Date_Text']

    df6["Inv_CN_DN_Date_Revised_Unique"] = df6['Inv_CN_DN_Date_Revised'].str.replace("-", ".")

    df6['Sheet_Name'] = ("CDNRA")

    df6['PAN_Number'] = df6["GSTIN_of_Supplier"].apply(lambda x: x[2:12:1])

    df6 = df6.replace(np.nan, "", regex=True)

    # Making a combined sheet with all merged

    print("We are Combining B2B, B2BA, CDNR & CDNRA in 1 sheet...Please wait..!")

    df8 = df3.append(df4)

    df9 = df8.append(df5)

    df10 = df9.append(df6)

    df10['PAN_Number'] = df10["GSTIN_of_Supplier"].apply(lambda x: x[2:12:1])

    df10 = df10.replace(np.nan, "", regex=True)

    df10["Ultimate_Unique"] = df10["Sheet_Name"] + "/" + df10["Supply_Attract_Reverse_Charge"] + df10[
        "GSTR_1_5_Filing_Status"] + "/" + df10["Unique_ID"]

    df10["PAN_3_Way_Key"] = np.where(df10["Sheet_Name"] == "B2BA",
                                     df10["PAN_Number"] + "/" + df10["Inv_CN_DN_Number_Revised"] + "/"
                                     + df10["Inv_CN_DN_Date_Revised_Unique"],
                                     df10["PAN_Number"] + "/" + df10["Inv_CN_DN_Number_Original"]
                                     + "/" + df10["Inv_CN_DN_Date_Text"])

    df10["PAN_2_Way_Key_PAN_InvNo"] = np.where(df10["Sheet_Name"] == "B2BA",
                                               df10["PAN_Number"] + "/" + df10["Inv_CN_DN_Number_Revised"]
                                               , df10["PAN_Number"] + "/" + df10["Inv_CN_DN_Number_Original"])

    df10["PAN_2_Way_Key_PAN_InvDt"] = np.where(df10["Sheet_Name"] == "B2BA",
                                               df10["PAN_Number"] + "/" + df10["Inv_CN_DN_Date_Revised_Unique"]
                                               , df10["PAN_Number"] + "/" + df10["Inv_CN_DN_Date_Text"])

    # maiking a sheet with person who did not file the GSTR 1

    df11 = df10[df10['GSTR_1_5_Filing_Status'] == "N"]

    df12 = df10[(df10['Supply_Attract_Reverse_Charge'] == "Y") & (df10['GSTR_1_5_Filing_Status'] == "Y")]

    df13 = df10[(df10['Supply_Attract_Reverse_Charge'] == "N") & (df10['GSTR_1_5_Filing_Status'] == "Y") & (
            df10['Total_Tax'] < 1)]

    df14 = df10[(df10['Supply_Attract_Reverse_Charge'] == "N") & (df10['GSTR_1_5_Filing_Status'] == "Y") & (
            df10['Total_Tax'] >= 1)]

    # saving the file with the name "Combined"

    extension = os.path.splitext(filepath)[1]
    filename = os.path.splitext(filepath)[0]
    pth = os.path.dirname(filepath)
    newfile = os.path.join(pth, filename + 'GSTR2A_all_combined' + extension)

    writer = pd.ExcelWriter(newfile, engine='openpyxl')

    print("Please wait.. we are creating different sheets and finalizing the file.....")

    df3.to_excel(writer, sheet_name="B2B")

    df4.to_excel(writer, sheet_name="B2BA")

    df5.to_excel(writer, sheet_name="CDNR")

    df6.to_excel(writer, sheet_name="CDNRA")

    titles = list(df10.columns)

    titles[0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], \
    titles[10], titles[11], titles[12], titles[13], titles[14], titles[15], titles[16], titles[17], titles[18], titles[
        19], titles[20], titles[21], titles[22], titles[23], titles[24], titles[25], titles[26], titles[27], titles[28], \
    titles[29], titles[30], titles[31], titles[32], titles[33], titles[34], titles[35] = titles[24], titles[28], titles[
        0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], titles[
                                                                                             10], titles[11], titles[
                                                                                             12], titles[13], titles[
                                                                                             26], titles[25], titles[
                                                                                             21], titles[27], titles[
                                                                                             14], titles[15], titles[
                                                                                             16], titles[17], titles[
                                                                                             18], titles[19], titles[
                                                                                             20], titles[22], titles[
                                                                                             23], titles[29], titles[
                                                                                             30], titles[31], titles[
                                                                                             32], titles[33], titles[
                                                                                             34], titles[35]

    df10[titles].to_excel(writer, sheet_name="All_Combined")

    titles = list(df11.columns)

    titles[0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], \
    titles[10], titles[11], titles[12], titles[13], titles[14], titles[15], titles[16], titles[17], titles[18], titles[
        19], titles[20], titles[21], titles[22], titles[23], titles[24], titles[25], titles[26], titles[27], titles[28], \
    titles[29], titles[30], titles[31], titles[32], titles[33], titles[34], titles[35] = titles[24], titles[28], titles[
        0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], titles[
                                                                                             10], titles[11], titles[
                                                                                             12], titles[13], titles[
                                                                                             26], titles[25], titles[
                                                                                             21], titles[27], titles[
                                                                                             14], titles[15], titles[
                                                                                             16], titles[17], titles[
                                                                                             18], titles[19], titles[
                                                                                             20], titles[22], titles[
                                                                                             23], titles[29], titles[
                                                                                             30], titles[31], titles[
                                                                                             32], titles[33], titles[
                                                                                             34], titles[35]

    df11[titles].to_excel(writer, sheet_name="GSTR_1_Not Filed")

    titles = list(df12.columns)

    titles[0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], \
    titles[10], titles[11], titles[12], titles[13], titles[14], titles[15], titles[16], titles[17], titles[18], titles[
        19], titles[20], titles[21], titles[22], titles[23], titles[24], titles[25], titles[26], titles[27], titles[28], \
    titles[29], titles[30], titles[31], titles[32], titles[33], titles[34], titles[35] = titles[24], titles[28], titles[
        0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], titles[
                                                                                             10], titles[11], titles[
                                                                                             12], titles[13], titles[
                                                                                             26], titles[25], titles[
                                                                                             21], titles[27], titles[
                                                                                             14], titles[15], titles[
                                                                                             16], titles[17], titles[
                                                                                             18], titles[19], titles[
                                                                                             20], titles[22], titles[
                                                                                             23], titles[29], titles[
                                                                                             30], titles[31], titles[
                                                                                             32], titles[33], titles[
                                                                                             34], titles[35]

    df12[titles].to_excel(writer, sheet_name="GSTR_Filed_RCM_Yes")

    titles = list(df13.columns)

    titles[0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], \
    titles[10], titles[11], titles[12], titles[13], titles[14], titles[15], titles[16], titles[17], titles[18], titles[
        19], titles[20], titles[21], titles[22], titles[23], titles[24], titles[25], titles[26], titles[27], titles[28], \
    titles[29], titles[30], titles[31], titles[32], titles[33], titles[34], titles[35] = titles[24], titles[28], titles[
        0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], titles[
                                                                                             10], titles[11], titles[
                                                                                             12], titles[13], titles[
                                                                                             26], titles[25], titles[
                                                                                             21], titles[27], titles[
                                                                                             14], titles[15], titles[
                                                                                             16], titles[17], titles[
                                                                                             18], titles[19], titles[
                                                                                             20], titles[22], titles[
                                                                                             23], titles[29], titles[
                                                                                             30], titles[31], titles[
                                                                                             32], titles[33], titles[
                                                                                             34], titles[35]

    df13[titles].to_excel(writer, sheet_name="Tax_Zero_Cases")

    titles = list(df14.columns)

    titles[0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], \
    titles[10], titles[11], titles[12], titles[13], titles[14], titles[15], titles[16], titles[17], titles[18], titles[
        19], titles[20], titles[21], titles[22], titles[23], titles[24], titles[25], titles[26], titles[27], titles[28], \
    titles[29], titles[30], titles[31], titles[32], titles[33], titles[34], titles[35] = titles[24], titles[28], titles[
        0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], titles[
                                                                                             10], titles[11], titles[
                                                                                             12], titles[13], titles[
                                                                                             26], titles[25], titles[
                                                                                             21], titles[27], titles[
                                                                                             14], titles[15], titles[
                                                                                             16], titles[17], titles[
                                                                                             18], titles[19], titles[
                                                                                             20], titles[22], titles[
                                                                                             23], titles[29], titles[
                                                                                             30], titles[31], titles[
                                                                                             32], titles[33], titles[
                                                                                             34], titles[35]

    df14[titles].to_excel(writer, sheet_name="Working_Cases")

    writer.save()

    print(f"All files have been combined and a single file named {newfile} has been created")

    return (writer)




def reco_itr_2a(files_itr,files_con2a,tol_limit=100):
    import numpy as np
    import openpyxl

    warnings.filterwarnings('ignore')



    print(f'The Consolidated GSTR2A file path is {files_con2a}')
    print(f'The ITR file path is {files_itr}')

    pth = os.path.dirname(str(files_con2a))

    fullpath1 = pth + "/" + "Workings.xlsx"

    writer = pd.ExcelWriter(fullpath1, engine='xlsxwriter', options={'strings_to_formulas': True})
    #




    fullpath1a = pth + "/" + "Summary.xlsx"
    writer1 = pd.ExcelWriter(fullpath1a, engine='xlsxwriter', options={'strings_to_formulas': True})

    df1 = pd.DataFrame()
    df1.to_excel(writer1, sheet_name="Summary", index=False)

    writer1.save()

    fullpath2 = fullpath1a.replace("/", "\\")  # this is a very useful command for defining the correct filepath

    wb = load_workbook(fullpath2)
    ws = wb["Summary"]

    ws["B2"].value = "SUMMARY OF THE RECONCILIATION OF GSTR2A Vs ITR"
    ws.merge_cells("B2:F2")
    ws["C4"].value = "GSTR2A"
    ws.merge_cells("C4:D4")
    ws["E4"].value = "Purchase Register"
    ws.merge_cells("E4:F4")

    ws["B4"].value = "Particulars"
    ws.merge_cells("B4:B5")
    ws["C5"].value = "Count"
    ws["D5"].value = "Tax Amount"
    ws["E5"].value = "Count"
    ws["F5"].value = "Tax Amount"

    ws["B7"].value = "Total cases in Original Files"
    ws["B8"].value = "Less: No GST Number in Purchase Register"
    ws["B9"].value = "Less: GSTR-1 Not filed cases"
    ws["B10"].value = "Less: GSTR-1 Filed but RCM cases "
    ws["B11"].value = "Less: No Invoice Number in Purchase Register"



    ws["B12"].value = "Net cases to be Matched"
    ws["B14"].value = "Matched with GST_INVNO_INVDATE_3_WAY"
    ws["B15"].value = "Matched with GST_INVNO_2_WAY"
    ws["B16"].value = "Matched with GST_INVDATE_2_WAY"

    ws["B18"].value = "Identified Possible Matches - Fuzzy Logic"

    ws["B20"].value = "Matched with PAN_INVNO_INVDATE_3_WAY"
    ws["B21"].value = "Matched with PAN_INVNO_2_WAY"
    ws["B22"].value = "Matched with PAN_INVDATE_2_WAY"

    ws["B24"].value = "Unmatched Cases"


    ws["B25"].value = "Unmatched Cases -PAN/GST not available in GSTR2A"
    ws["B26"].value = "Unmatched Cases with Invalid GSTIN"


    ws["B27"].value = "Unmatched cases with Invalid Invoice Number"

    ws["B28"].value = "Other Unmatched Cases"

    ws["B30"].value = "Check"

    # setting the tolerance limit for matching in Rupees

    tol_limit = int(tol_limit)

    print(f"The tolerance limit is set to {tol_limit}")

    ws["F1"].value = f"Tolerance Limit was {tol_limit}"

    gstr2a = pd.read_excel(files_con2a, sheet_name="Main_2A_Format",dtype={"Inv_CN_DN_Number_Final":str, "Inv_CN_DN_Date_Text":str, "Total_Tax":int})

    try:
        gstr2a['Inv_CN_DN_Number_Finall'] = gstr2a["Inv_CN_DN_Number_Final"].apply(lambda x: x.lower(str()))
    except:
        gstr2a['Inv_CN_DN_Number_Finall'] = gstr2a["Inv_CN_DN_Number_Final"]

    gstr2a['GST_INVNO_INVDATE_3_WAY'] = gstr2a['GSTIN_of_Supplier'] + "/" + gstr2a['Inv_CN_DN_Number_Finall'] + "/" + gstr2a['Inv_CN_DN_Date_Text']

    gstr2a['GST_INVNO_2_WAY'] = gstr2a['GSTIN_of_Supplier'] + "/" + gstr2a['Inv_CN_DN_Number_Finall']

    gstr2a['GST_INVDATE_2_WAY'] = gstr2a['GSTIN_of_Supplier'] + "/" + gstr2a['Inv_CN_DN_Date_Text']

    
    gstr2a['PAN_Number'] = gstr2a["GSTIN_of_Supplier"].apply(lambda x:extract_pan(x))
    
    # the PAN number matches will be used as possible matches

    gstr2a['PAN_INVNO_INVDATE_3_WAY'] = gstr2a['PAN_Number'] + "/" + gstr2a['Inv_CN_DN_Number_Finall'] + "/" + \
                                        gstr2a['Inv_CN_DN_Date_Text']

    gstr2a['PAN_INVNO_2_WAY'] = gstr2a['PAN_Number'] + "/" + gstr2a['Inv_CN_DN_Number_Finall']

    gstr2a['PAN_INVDATE_2_WAY'] = gstr2a['PAN_Number'] + "/" + gstr2a['Inv_CN_DN_Date_Text']


    itr = pd.read_excel(files_itr, sheet_name="Main_ITR_Format",dtype={"Invoice_Number":str, "Invoice_Date_Text":str,"Total_Tax":int})

    try:
        itr["Invoice_Numberl"] = itr["Invoice_Number"].apply(lambda x: x.lower(str()))
    except:
        itr["Invoice_Numberl"] = itr["Invoice_Number"]

    itr["GST_INVNO_INVDATE_3_WAY"] = itr["Vendor_GST_REG"] + "/" + itr["Invoice_Numberl"] + "/" + itr[
        "Invoice_Date_Text"]

    itr["GST_INVNO_2_WAY"] = itr["Vendor_GST_REG"] + "/" + itr["Invoice_Numberl"]

    itr["GST_INVDATE_2_WAY"] = itr["Vendor_GST_REG"] + "/" + itr["Invoice_Date_Text"]

   
    itr["PAN_Number"] = itr["Vendor_GST_REG"].apply(lambda x:extract_pan(x))
    


    # the PAN number matches will be used as possible matches

    itr["PAN_INVNO_INVDATE_3_WAY"] = itr["PAN_Number"] + "/" + itr["Invoice_Numberl"] + "/" + itr["Invoice_Date_Text"]

    itr["PAN_INVNO_2_WAY"] = itr["PAN_Number"] + "/" + itr["Invoice_Numberl"]

    itr["PAN_INVDATE_2_WAY"] = itr["PAN_Number"] + "/" + itr["Invoice_Date_Text"]

    ws["C7"].value = list(gstr2a.shape)[0]
    ws["D7"].value = sum(gstr2a["Total_Tax"])
    ws["E7"].value = list(itr.shape)[0]
    ws["F7"].value = sum(itr["Total_Tax"])



    #data Cleaning for GSTr2A:

    try:
        
        gstr2a_not_filed=gstr2a[gstr2a['GSTR_1_5_Filing_Status'] == "N"]

    except:

        pass


    try:
        gstr2a_rcm=gstr2a[(gstr2a['Supply_Attract_Reverse_Charge'] == "Y") & (gstr2a['GSTR_1_5_Filing_Status'] == "Y")]
    except:
        pass


    try:
        gstr2a_work=gstr2a[(gstr2a['Supply_Attract_Reverse_Charge'] != "Y") & (gstr2a['GSTR_1_5_Filing_Status'] == "Y")]
    except:
        gstr2a_work=gstr_2a
        

    ws["C9"].value = len(gstr2a_not_filed["GSTIN_of_Supplier"])
    ws["D9"].value = sum(gstr2a_not_filed["Total_Tax"])
    ws["C10"].value = len(gstr2a_rcm["GSTIN_of_Supplier"])
    ws["D10"].value = sum(gstr2a_rcm["Total_Tax"])

    ws["C12"].value = len(gstr2a_work["GSTIN_of_Supplier"])
    ws["D12"].value = sum(gstr2a_work["Total_Tax"])


    #data cleaing  for ITR as of now is only Blank GST Reg No and Blank Invoice Number.
    #So, net case to be matched will equal to Total cases

    mask=itr["Vendor_GST_REG"].isnull()

    try:
        itr_nogst=itr[mask]
    except:
        pass


    itr_gst=itr[~mask]

    mask2=itr_gst["Invoice_Numberl"].isnull()

    try:
        itr_noinvno=itr_gst[mask2]
    except:
        pass

    itr_work=itr_gst[~mask2]


    ws["E8"].value = list(itr_nogst.shape)[0]
    ws["F8"].value = sum(itr_nogst["Total_Tax"])

    ws["E11"].value = list(itr_noinvno.shape)[0]
    ws["F11"].value = sum(itr_noinvno["Total_Tax"])




    ws["E12"].value = list(itr_work.shape)[0]
    ws["F12"].value = sum(itr_work["Total_Tax"])




    # First Cut Matching : Here we will try to do that Matching based on 3 way i.e GST No, Inv No & Inv Date being same in ITR & GSTR2A



    gstr2a_pivot = pd.pivot_table(gstr2a_work, values="Total_Tax", index=["GST_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

    itr_pivot = pd.pivot_table(itr_work, values="Total_Tax", index=["GST_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="GST_INVNO_INVDATE_3_WAY", right_on="GST_INVNO_INVDATE_3_WAY",
                                  how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks"].values == "Exact Match within Tolerance"

    match_3_way_list = compared[mask_1]["GST_INVNO_INVDATE_3_WAY"].values

    mask_1a = gstr2a["GST_INVNO_INVDATE_3_WAY"].isin(match_3_way_list)  # returns a Boolean Array

    mask_1b = itr["GST_INVNO_INVDATE_3_WAY"].isin(match_3_way_list)  # returns a boolean array

    matched_gstr2a_3way = gstr2a[mask_1a]
    matched_gstr2a_3way["Matching Category"] = "3 Way matching GST + Inv No+ Inv Date"
    matched_itr_3way = itr[mask_1b]
    matched_itr_3way["Matching Category"] = "3 Way matching GST + Inv No+ Inv Date"

    ws["C14"].value = len(matched_gstr2a_3way["GST_INVNO_INVDATE_3_WAY"])
    ws["D14"].value = sum(matched_gstr2a_3way["Total_Tax"])
    ws["E14"].value = len(matched_itr_3way["GST_INVNO_INVDATE_3_WAY"])
    ws["F14"].value = sum(matched_itr_3way["Total_Tax"])

    bal_gstr2a_1cut = gstr2a_work[~mask_1a]
    bal_itr_1cut = itr_work[~mask_1b]

    # Second Cut Matching : Here we will try to do that Matching based on 2 way i.e GST No & Inv No

    gstr2a_pivot = pd.pivot_table(bal_gstr2a_1cut, values="Total_Tax", index=["GST_INVNO_2_WAY"], aggfunc=np.sum)

    itr_pivot = pd.pivot_table(bal_itr_1cut, values="Total_Tax", index=["GST_INVNO_2_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="GST_INVNO_2_WAY", right_on="GST_INVNO_2_WAY", how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks"].values == "Exact Match within Tolerance"

    match_2_way_list1 = compared[mask_1]["GST_INVNO_2_WAY"].values

    mask_1a = bal_gstr2a_1cut["GST_INVNO_2_WAY"].isin(match_2_way_list1)  # returns a Boolean Array

    mask_1b = bal_itr_1cut["GST_INVNO_2_WAY"].isin(match_2_way_list1)  # returns a boolean array

    matched_gstr2a_2way1 = bal_gstr2a_1cut[mask_1a]
    matched_itr_2way1 = bal_itr_1cut[mask_1b]

    matched_gstr2a_2way1["Matching Category"] = "2 Way matching GST + Inv No"
    matched_itr_2way1["Matching Category"] = "2 Way matching GST + Inv No"

    ws["C15"].value = len(matched_gstr2a_2way1["GST_INVNO_2_WAY"])
    ws["D15"].value = sum(matched_gstr2a_2way1["Total_Tax"])
    ws["E15"].value = len(matched_itr_2way1["GST_INVNO_2_WAY"])
    ws["F15"].value = sum(matched_itr_2way1["Total_Tax"])

    bal_gstr2a_2cut = bal_gstr2a_1cut[~mask_1a]
    bal_itr_2cut = bal_itr_1cut[~mask_1b]

    # Third Cut Matching : Here we will try to do that Matching based on 2 way i.e GST No & Inv Date

    gstr2a_pivot = pd.pivot_table(bal_gstr2a_2cut, values="Total_Tax", index=["GST_INVDATE_2_WAY"], aggfunc=np.sum)

    itr_pivot = pd.pivot_table(bal_itr_2cut, values="Total_Tax", index=["GST_INVDATE_2_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="GST_INVDATE_2_WAY", right_on="GST_INVDATE_2_WAY", how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks"].values == "Exact Match within Tolerance"

    match_2_way_list2 = compared[mask_1]["GST_INVDATE_2_WAY"].values

    mask_1a = bal_gstr2a_2cut["GST_INVDATE_2_WAY"].isin(match_2_way_list2)  # returns a Boolean Array

    mask_1b = bal_itr_2cut["GST_INVDATE_2_WAY"].isin(match_2_way_list2)  # returns a boolean array

    matched_gstr2a_2way2 = bal_gstr2a_2cut[mask_1a]
    matched_itr_2way2 = bal_itr_2cut[mask_1b]

    matched_gstr2a_2way2["Matching Category"] = "2 Way matching GST + Inv Date"
    matched_itr_2way2["Matching Category"] = "2 Way matching GST + Inv Date"

    ws["C16"].value = len(matched_gstr2a_2way2["GST_INVDATE_2_WAY"])
    ws["D16"].value = sum(matched_gstr2a_2way2["Total_Tax"])
    ws["E16"].value = len(matched_itr_2way2["GST_INVDATE_2_WAY"])
    ws["F16"].value = sum(matched_itr_2way2["Total_Tax"])

    bal_gstr2a_3cut = bal_gstr2a_2cut[~mask_1a]
    bal_itr_3cut = bal_itr_2cut[~mask_1b]

    print(f"The 3 way matching using GST is done.... Now ,we are doing the mtching using PAN..Please wait...!!")




    #after the 3 cut matching, now we try to find out the Possible matches in form of PAN matching and upper /lower case matching
    # Fourth Cut Matching : Here we will try to do that Matching based on 3 way With PAN No Inv NO and Inv Date

    gstr2a_pivot = pd.pivot_table(bal_gstr2a_3cut, values="Total_Tax", index=["PAN_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

    itr_pivot = pd.pivot_table(bal_itr_3cut, values="Total_Tax", index=["PAN_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="PAN_INVNO_INVDATE_3_WAY", right_on="PAN_INVNO_INVDATE_3_WAY", how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks"].values == "Exact Match within Tolerance"

    match_3_way_list2 = compared[mask_1]["PAN_INVNO_INVDATE_3_WAY"].values

    mask_1a = bal_gstr2a_3cut["PAN_INVNO_INVDATE_3_WAY"].isin(match_3_way_list2)  # returns a Boolean Array

    mask_1b = bal_itr_3cut["PAN_INVNO_INVDATE_3_WAY"].isin(match_3_way_list2)  # returns a boolean array

    matched_gstr2a_3way2 = bal_gstr2a_3cut[mask_1a]
    matched_itr_3way2 = bal_itr_3cut[mask_1b]

    matched_gstr2a_3way2["Matching Category"] = "3 Way matching PAN + Inv No+ Inv Date"
    matched_itr_3way2["Matching Category"] = "3 Way matching PAN + Inv No + Inv Date"

    ws["C20"].value = len(matched_gstr2a_3way2["PAN_INVNO_INVDATE_3_WAY"])
    ws["D20"].value = sum(matched_gstr2a_3way2["Total_Tax"])
    ws["E20"].value = len(matched_itr_3way2["PAN_INVNO_INVDATE_3_WAY"])
    ws["F20"].value = sum(matched_itr_3way2["Total_Tax"])

    bal_gstr2a_4cut = bal_gstr2a_3cut[~mask_1a]
    bal_itr_4cut = bal_itr_3cut[~mask_1b]

    # Fifth Cut Matching : Here we will try to do that Matching based on 3 way With PAN No Inv NO and Inv Date

    gstr2a_pivot = pd.pivot_table(bal_gstr2a_4cut, values="Total_Tax", index=["PAN_INVNO_2_WAY"],
                                  aggfunc=np.sum)

    itr_pivot = pd.pivot_table(bal_itr_4cut, values="Total_Tax", index=["PAN_INVNO_2_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="PAN_INVNO_2_WAY", right_on="PAN_INVNO_2_WAY",
                                  how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks"].values == "Exact Match within Tolerance"

    match_2_way_list3 = compared[mask_1]["PAN_INVNO_2_WAY"].values

    mask_1a = bal_gstr2a_4cut["PAN_INVNO_2_WAY"].isin(match_2_way_list3)  # returns a Boolean Array

    mask_1b = bal_itr_4cut["PAN_INVNO_2_WAY"].isin(match_2_way_list3)  # returns a boolean array

    matched_gstr2a_2way3 = bal_gstr2a_4cut[mask_1a]
    matched_itr_2way3 = bal_itr_4cut[mask_1b]

    matched_gstr2a_2way3["Matching Category"] = "2 Way matching PAN + Inv No"
    matched_itr_2way3["Matching Category"] = "2 Way matching PAN + Inv No "

    ws["C21"].value = len(matched_gstr2a_2way3["PAN_INVNO_2_WAY"])
    ws["D21"].value = sum(matched_gstr2a_2way3["Total_Tax"])
    ws["E21"].value = len(matched_itr_2way3["PAN_INVNO_2_WAY"])
    ws["F21"].value = sum(matched_itr_2way3["Total_Tax"])

    bal_gstr2a_5cut = bal_gstr2a_4cut[~mask_1a]
    bal_itr_5cut = bal_itr_4cut[~mask_1b]



    # Sixth Cut Matching : Here we will try to do that Matching based on 3 way With PAN No Inv NO and Inv Date

    gstr2a_pivot = pd.pivot_table(bal_gstr2a_5cut, values="Total_Tax", index=["PAN_INVDATE_2_WAY"],
                                  aggfunc=np.sum)

    itr_pivot = pd.pivot_table(bal_itr_4cut, values="Total_Tax", index=["PAN_INVDATE_2_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="PAN_INVDATE_2_WAY", right_on="PAN_INVDATE_2_WAY",
                                  how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks"].values == "Exact Match within Tolerance"

    match_2_way_list4 = compared[mask_1]["PAN_INVDATE_2_WAY"].values

    mask_1a = bal_gstr2a_5cut["PAN_INVDATE_2_WAY"].isin(match_2_way_list4)  # returns a Boolean Array

    mask_1b = bal_itr_5cut["PAN_INVDATE_2_WAY"].isin(match_2_way_list4)  # returns a boolean array

    matched_gstr2a_2way4 = bal_gstr2a_5cut[mask_1a]
    matched_itr_2way4 = bal_itr_5cut[mask_1b]

    matched_gstr2a_2way4["Matching Category"] = "2 Way matching PAN + Inv Date"
    matched_itr_2way4["Matching Category"] = "2 Way matching PAN + Inv Date "

    ws["C22"].value = len(matched_gstr2a_2way4["PAN_INVDATE_2_WAY"])
    ws["D22"].value = sum(matched_gstr2a_2way4["Total_Tax"])
    ws["E22"].value = len(matched_itr_2way4["PAN_INVDATE_2_WAY"])
    ws["F22"].value = sum(matched_itr_2way4["Total_Tax"])

    bal_gstr2a_6cut = bal_gstr2a_5cut[~mask_1a]
    bal_itr_6cut = bal_itr_5cut[~mask_1b]

    
    #NOw, after all matching, we are further analyzing the et Unmatched Cases


    print(f"Analyzing the Unmatched cases of ITR.... Please wait..!")

    #First, we check whether the PAN we are searching is present in GSTr2A at all or not. If not present then
    #we identify it separately. These have absolute no chaces of matching


    pan_itr=list(set(list(bal_itr_6cut["PAN_Number"].values)))
    pan_2a=bal_gstr2a_6cut["PAN_Number"].values

    maskpan=bal_itr_6cut["PAN_Number"].isin(pan_2a)

    bal_itr_6cut1=bal_itr_6cut[maskpan]

    unmatched_itr1=bal_itr_6cut[~maskpan]

    unmatched_itr1["Remarks"]="PAN/GST not available in GSTR2A"

    ws["E25"].value = len(unmatched_itr1["Remarks"])
    ws["F25"].value = sum(unmatched_itr1["Total_Tax"])


    #Second, we will see the CheckSUm Digit of the GST Number. Whether the last charater which is acheck sum is matching or not
    #this is also very crucials, as If GSTIN is invalid, there is no point of matching



    bal_itr_6cut1["GSTN Status"]=bal_itr_6cut1["Vendor_GST_REG"].apply(lambda x:gstchecksum(x))

    mask1=bal_itr_6cut1["GSTN Status"].values=="Check Sum MATCH"


    bal_itr_6cut2=bal_itr_6cut1[mask1]

    unmatched_itr2=bal_itr_6cut1[~mask1]


    unmatched_itr2["Remarks"]="GST Number Check Sum Incorrect"



    ws["E26"].value = len(unmatched_itr2["Remarks"])
    ws["F26"].value = sum(unmatched_itr2["Total_Tax"])


    #Third, we will be checking the Invoice Number check
    #if Invoice Number exceeds 16 digits , then we will be marking these seprately as no chaces of matching

    bal_itr_6cut2["Invoice No Check"]=bal_itr_6cut2["Invoice_Number"].apply(lambda x:gstinvcheck(x))

    mask2=bal_itr_6cut2["Invoice No Check"].values=="Invoice Number Valid"


    bal_itr_6cut3=bal_itr_6cut2[mask2]
    unmatched_itr3=bal_itr_6cut2[~mask2]

    unmatched_itr3["Remarks"]="Invoice No length exceed 16 digit"

    

    ws["E27"].value = len(unmatched_itr3["Remarks"])
    ws["F27"].value = sum(unmatched_itr3["Total_Tax"])


    #hre we will ttry to do the fuzzy matching of the Invoice Number with the GSTR2A

    print("Trying to do some Fuzzy matches in GSTR2A and ITR. Please wait....!!")


    from difflib import SequenceMatcher , get_close_matches

    cant_match=[]
    matches_itr=[]
    matches_gstr2ai=[]

    df=list(set(list(bal_itr_6cut3["PAN_Number"].values)))

    for i in df:
    #     Here the i variable is storing the PAN number each time the loop runs
        itr_balinv=bal_itr_6cut3[bal_itr_6cut3["PAN_Number"].values==i]["Invoice_Number"].values
        
        # print(f"This is ITR Invoice of {i}")
        # print(itr_balinv)
        
        gstr2a_balinv=(bal_gstr2a_6cut[bal_gstr2a_6cut["PAN_Number"].values==i]["Inv_CN_DN_Number_Finall"].values).tolist()
        # print(f"This is GSTR2A Invoice of {i}")
        # print(gstr2a_balinv)
        

        
        zipped=zip(itr_balinv,gstr2a_balinv)
        
        
        if len(gstr2a_balinv)==0:
            cant_match.append(itr_balinv)
            
        
            
        else:
               
            for inv in itr_balinv:
                
                matches_gstr2a=get_close_matches(inv,gstr2a_balinv,n=1,cutoff=0.90)
                
                if len(matches_gstr2a)==1:
                    
                    # print(f"this is inv{inv}")
                    matches_itr.append(inv)
                    matches_gstr2ai.append(matches_gstr2a[0])
                    
                    try:
                        gstr2a_balinv.remove(matches_gstr2a[0])
                    except:
                        continue
                else:
                    
                    continue
            
          
            
            cant_match.append(list(set(itr_balinv)-set(matches_itr)))
        

    mask1a=bal_itr_6cut3["Invoice_Number"].isin(matches_itr)

    mask1b=bal_gstr2a_6cut["Inv_CN_DN_Number_Finall"].isin(matches_gstr2ai)



    prob_itr_match=bal_itr_6cut3[mask1a]

    prob_gstr2a_match=bal_gstr2a_6cut[mask1b]

    prob_gstr2a_match["Matching Category"] = "Probable Match- Fuzzy Logic"
    prob_itr_match["Matching Category"] = "Probable Match- Fuzzy Logic"



    bal_itr_6cut4=bal_itr_6cut3[~mask1a]

    bal_gstr2a_7cut=bal_gstr2a_6cut[~mask1b]

    bal_itr_6cut4["Remarks"]="These Cases are Not Matching"


    ws["C18"].value = len(prob_gstr2a_match["Inv_CN_DN_Number_Finall"])
    ws["D18"].value = sum(prob_gstr2a_match["Total_Tax"])
    ws["E18"].value = len(prob_itr_match["Invoice_Number"])
    ws["F18"].value = sum(prob_itr_match["Total_Tax"])


    print(f"Matchig is done...Creating the 2 Files for you. Summary.xlsx & Working.xlsx")



    #now, we will be merging all these Unmatched cases of itr and final balance cut ITR

    bal_itr_7cut=pd.concat([unmatched_itr1,unmatched_itr2,unmatched_itr3,bal_itr_6cut4])


    gstr2a.to_excel(writer, sheet_name='Orignal GSTR2A', index=False)

    itr.to_excel(writer, sheet_name='Original ITR', index=False)

    all_matched_2a = pd.concat([matched_gstr2a_3way, matched_gstr2a_2way1, matched_gstr2a_2way2,matched_gstr2a_3way2, matched_gstr2a_2way3,matched_gstr2a_2way4,prob_gstr2a_match], ignore_index=True)

    all_matched_itr = pd.concat([matched_itr_3way, matched_itr_2way1, matched_itr_2way2,matched_itr_3way2,matched_itr_2way3, matched_itr_2way4,prob_itr_match], ignore_index=True)

    all_matched_2a.to_excel(writer, sheet_name='Matched_GSTR2A', index=False)

    all_matched_itr.to_excel(writer, sheet_name='Matched_ITR', index=False)

    bal_gstr2a_7cut.to_excel(writer, sheet_name='Unmatched_GSTR2A', index=False)

    bal_itr_7cut.to_excel(writer, sheet_name='Unmatched_ITR', index=False)


    ws["C28"].value = len(bal_gstr2a_7cut["GST_INVDATE_2_WAY"])
    ws["D28"].value = sum(bal_gstr2a_7cut["Total_Tax"])
    ws["E28"].value = len(bal_itr_6cut4["Remarks"])
    ws["F28"].value = sum(bal_itr_6cut4["Total_Tax"])



    writer.save()

    print("Success! ")



    wb.save(fullpath2)
    writer.save()

    print(f'Matching has been done and saved in below path \n {fullpath2}\n ')



"""
So, this is a function to download the GSTR2A & ITR format and also instructions to use this utility

User have an option to provide a path in which they want to store the Formats. If path is not provided , the format will be download in Current Workig Directory
"""

def download(pth=os.getcwd()):
    import pandas as pd
    import numpy as np
    import openpyxl


    # pth = os.getcwd()

    fullpath1 = pth + "\\" + "Formats.xlsx"
    print(fullpath1)

    writer = pd.ExcelWriter(fullpath1, engine='xlsxwriter', options={'strings_to_formulas': True})

    dict1 = {"Vendor_GST_REG": ["Mandatory"], "Vendor_Name": ["Optional"], "Invoice_Number": ["Mandatory"],
             "Invoice_Date_Text": ["Mandatory"], "Total_Tax": ["Mandatory"], "IGST": ["Optional"], "SGST": ["Optional"],
             "CGST": ["Optional"], "UTGST": ["Optional"],"User Defined1":["Optional"],"User Defined2":["Optional"],"User Defined3":["Optional"],"User Defined4":["Optional"]}

    df1 = pd.DataFrame(dict1)
    df1.to_excel(writer, sheet_name="Main_ITR_Format", index=False)

    dict2={"GSTIN_of_Supplier":["Mandatory"],"Inv_CN_DN_Number_Final":["Mandatory"],"Legal_Name_Of Supplier":["Optional"],"Inv_CN_DN_Date_Text":["Mandatory"],"Total_Tax":["Mandatory"],"GSTR_1_5_Filing_Status":["Mandatory"],"Supply_Attract_Reverse_Charge":["Mandatory"],"IGST":["Optional"],"SGST":["Optional"],"CGST":["Optional"],"UTGST":["Optional"],"User Defined1":["Optional"],"User Defined2":["Optional"],"User Defined3":["Optional"],"User Defined4":["Optional"]}
    df2 = pd.DataFrame(dict2)
    df2.to_excel(writer, sheet_name="Main_2A_Format", index=False)

    dict3={"Things to ensure before Running the Program":[" ","Keep GSTR2A and ITR in different Folders ","In that Folder , there should not be any other excel files",
    "Format of the ITR & GSTR2A can be as per the user ","However, below points to be taken care","The name of the Sheet having the  ITR should be Main_ITR_Format ",
    "The name of the Sheet having the Consolidated GSTR2A should be Main_2A_Format "," ","There are 6 Mandatory columns in both GSTR2A and ITR ","The Name to be assigned to these 6 Mandatory columns must be same as in the format",
    "In ITR , Mandatory columns are Vendor_GST_REG , Invoice_Number, Invoice_Date_Text, Total_Tax","Even the upper and Lower case should be same as in the Format",
    "In GSTR2A , Mandatory columns are GSTIN_of_Supplier , Inv_CN_DN_Number_Final, Inv_CN_DN_Date_Text, Total_Tax,Supply_Attract_Reverse_Charge,GSTR_1_5_Filing_Status","Take care of the Upper and Lower case and special Character",
    "The sequence of the Columns is not relevant. User can maintain the Sequence of the columns as per his own convinience","The ITR or GSTR2A file can also have multiple other sheets as per need of user , but relevant data for matching should be in one sheet only .",
    "But it has to be ensured that the main sheet in the GSTR2A and ITR has exactly the same namee as mentioned in the Format", "  "," ","For any issues in running the Code, send your issues to efficientcorporates.info@gmail.com",
    "For more such Automation Videos, Follow YouTube Channel Efficient Corporates"]}
    
    df3 = pd.DataFrame(dict3)
    df3.to_excel(writer, sheet_name="Important_Checklist", index=False)


    writer.save()

    print(f'The Formats have been saved in below path \n {fullpath1}\n ')
