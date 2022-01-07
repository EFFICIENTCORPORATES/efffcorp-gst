import pandas as pd
import numpy as np
import os
import glob
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from shutil import copyfile
import datetime
import warnings
from difflib import SequenceMatcher , get_close_matches

warnings.filterwarnings('ignore')



def gstchecksum(gst_no):


    """
    This function gstchecksum will check the last digit of the gst number and return whether the Check sum matches or not

    Output will be given as "Check Sum MATCH" or "Check Sum MISMATCH"

    :param gst_no: This is the only argument that needs to be given. It is a mandatory Argument.The argument must be 15 digit long

    :type gst_no: This parameter must be a string and must be 15 digit long

    :return : The function will return only one of two values a) Check Sum MATCH & b) Check Sum MISMATCH

    :raises: There are two errors that will be raised by the Function 
            1. Type Error: If the parameter entered is not a string, then this error is raised
            2. Exception: If the parameter entered is not 15 digit long, then Exception is raised

    :see also: To know how the GST Ceck sum is calculated , see the alogorithm behind the last digit



    """

    charlist=[char for char in gst_no.upper()]

    a=1
    cumhash=[]

    if not type(gst_no) is str:
        raise TypeError("Only strings are allowed")


    try:
        if len(gst_no)<15:
            raise Exception("Please ensure that the input is 15 digit long")
    except:
        pass


    for i in charlist[0:14:1]:
        
        if a % 2==0:
            multiplier=2
        else:
            multiplier=1

        if i.isdigit():
            intvalue=int(i)
            prod=intvalue*multiplier
            quotient=prod//36
            remain=prod%36
            hash=quotient+remain
            
        else:
            intvalue=ord(i)-55   
            prod=intvalue*multiplier
            quotient=prod//36
            remain=prod%36
            hash=quotient+remain
            

        a=a+1

        cumhash.append(hash)

    hashsum=(sum(cumhash))

    remain=hashsum%36

    checksum=36-remain

    if checksum<10:
        finalchk=int(checksum)
    else:
        finalchk=chr(checksum+55)
   
    
    lastchr=(gst_no[len(gst_no)-1])
    
    if lastchr.isdigit():
        lastchr=int(lastchr)
    else:
        lastchr=str(lastchr)
        
        
    if finalchk==lastchr:
        result="Check Sum MATCH"
    else:
        result="Check Sum MISMATCH"
    
    return (result)
     




def getgstcheck(number):


    """
    This function getgstcheck will give the last digit of the gst number

    Output will be given as the last digit which should be as per the given 14 digit number

    :param number: This is the only argument that needs to be given. It is a mandatory Argument.The argument must beat least 14 digit long

    :type gst_no: This parameter must be a string and must be at least digit long

    :return : The function will return the correct last digit of the given gst number

    :raises: There are two errors that will be raised by the Function 
            1. Type Error: If the parameter entered is not a string, then this error is raised
            2. Exception: If the parameter entered is not at least 14 digit long, then Exception is raised

    :see also: To know how the GST Ceck sum is calculated , see the alogorithm behind the last digit



    """

    charlist=[char for char in number.upper()]

    a=1
    cumhash=[]

    if not type(number) is str:
        raise TypeError("Only strings are allowed")


    if len(str((number)))<14:
        
        
        print ("Please ensure that the input is at least 14 digit long")
        
        pass

    else:
        
        pass
        


    for i in charlist[0:14:1]:
        
        if a % 2==0:
            multiplier=2
        else:
            multiplier=1

        if i.isdigit():
            intvalue=int(i)
            prod=intvalue*multiplier
            quotient=prod//36
            remain=prod%36
            hash=quotient+remain
            
        else:
            intvalue=ord(i)-55   
            prod=intvalue*multiplier
            quotient=prod//36
            remain=prod%36
            hash=quotient+remain
            

        a=a+1

        cumhash.append(hash)

    hashsum=(sum(cumhash))

    remain=hashsum%36

    checksum=36-remain

    if checksum<10:
        finalchk=int(checksum)
    else:
        finalchk=chr(checksum+55)
   
    
    
    return (finalchk)


def gstinvcheck(a):

    """
    This function will check whether the invoice number entered is correct or not.

    As per GST rules, the Invoice number must be maximum 15 digit long

    :param a: this must be the GST  Invoice number 

    :param a type: The Type of parameter must be a string. However, in the functionit is converting any parameter into a string through str() method

    :return : it return one of 2 output  a) Invoice Number Valid or b) Invoice Number Invalid



    """

    try:
        length=len(str(a))
    except:
        length=0    


    if length<=16:
        status="Invoice Number Valid"
    else:
        status="Invoice Number Invalid"
        
    return(status)





def extract_pan(gst_no):



    """
    This function will extract the PAN number from the provided GST_No

    :param gst_no: This function requires only one parameter. ie the GST No

    :param gst_no type: The type of the parameter must be a string

    :return :The function will return a string which is the PAN Number

    :SeeAlso : The PAN number is the 3rd Character to 12th Character of the GST Number



    """
    try:
        if not type(gst_no) is str:
            raise TypeError("Only strings are allowed")

        else:
            pass
            

    except:
        pass
        


    try:
        if len(gst_no)<15:
            raise Exception("Please ensure that the input is 15 digit long")

        else:
            pass
            
    except:
        pass

    

    try:
        pan_num=gst_no[2:12:1]
    except:
        pan_num=gst_no
    
    return(pan_num)



def gstr2a_merge(filepath):
    """
    This is a super useful Function for merging all the GSTR2A files kept in a folder.


    :Param filepath: The function takes only one parameter which is the Compelte File Path to any one GSTR2A file in that folder

    :Param type : The type of the argument should be the complete path to the excel file of the GSTR2A , till the extension 

    :return : The function will return a merged Excel File by merging all the GSTR2A Sheets i.e the B2B, B2BA, CDNR, CDNRA

    :Warnings:  Please note that only one argument needs to be passed
                
                The Function will auto read all other files which are there in the folder

                Please ensure that the particular folder only has the GSTR2A files that you want to combine. 

                There should not be any other files as this function will read all the files in that folder whether or not they are GSTR2A files

    :See Also: Just a snippet of how this function works

                The code will first loop through all the B2B files of the folder and then B2B A, then CDNR and then CDNRA.

                After looping through all the files, this will first make 4 different sheets of these 4 types

                Then it will make a All Combined Files by merging all the B2B, B2BA, CDNR and CDNRA

                Also, as an additional analysis, this will make 3 different sheets a) RCM Cases b) GSTR-1 Filing Status as No c) tax Zero cases

                NOw, as per the GST Act, and Rule 36, for claiming the ITC, GSTR -1 Fling status to be Y and RCM should be N



    """

    import pandas as pd
    import glob
    import os


    pth = os.path.dirname(filepath)

    filenames = glob.glob(pth + "/*.xlsx")

    warnings.filterwarnings('ignore')

    

    cum_size = 0

    for file in filenames:
        size = os.path.getsize(file)

        cum_size = cum_size + size

        if size > 31457280:
            print("Please upload a smaller file size. Maximum limit is 30 mb.")

        elif cum_size > 314572800:
            print("Combined File size for all the file is more than 300 mb. Please use smaller files")
            break
        else:
            pass

    # A. iterate through each file to append it one below the other

    # A.1 : This will iterate through the B2B file


    print(f"The files that will be combined are \n {filenames}")


    print("We are working on B2B sheet of all the monthly GSTR2A..Please wait...")


    df2 = pd.DataFrame()

    for files in filenames:
        df = pd.read_excel(files, sheet_name=1)

        df1 = df.drop([0, 1, 2, 3, 4], axis=0)

        df1 = df1.dropna(how='all')

        df1['File_name'] = files

        df2 = df2.append(df1)

    # this is used for deleting all the rows which are totally blank

    df3 = df2

    # this is used for renaming the names of the columns

    df3.rename(columns={'Goods and Services Tax  - GSTR 2A': 'GSTIN_of_Supplier'}, inplace=True)
    df3.rename(columns={'Unnamed: 1': 'Legal_Name_Of Supplier'}, inplace=True)
    df3.rename(columns={'Unnamed: 2': 'Inv_CN_DN_Number_Original'}, inplace=True)
    df3.rename(columns={'Unnamed: 3': 'Inv_CN_DN_Type_Original'}, inplace=True)
    df3.rename(columns={'Unnamed: 4': 'Inv_CN_DN_Date_Original'}, inplace=True)
    df3.rename(columns={'Unnamed: 5': 'Inv_CN_DN_Value_Original'}, inplace=True)
    df3.rename(columns={'Unnamed: 6': 'Place_Of_Supply'}, inplace=True)
    df3.rename(columns={'Unnamed: 7': 'Supply_Attract_Reverse_Charge'}, inplace=True)
    df3.rename(columns={'Unnamed: 8': 'GST_Rate'}, inplace=True)
    df3.rename(columns={'Unnamed: 9': 'Taxable_Value_Rs'}, inplace=True)
    df3.rename(columns={'Unnamed: 10': 'IGST_Rs'}, inplace=True)
    df3.rename(columns={'Unnamed: 11': 'CGST_Rs'}, inplace=True)
    df3.rename(columns={'Unnamed: 12': 'SGST_Rs'}, inplace=True)
    df3.rename(columns={'Unnamed: 13': 'Cess'}, inplace=True)
    df3.rename(columns={'Unnamed: 14': 'GSTR_1_5_Filing_Status'}, inplace=True)
    df3.rename(columns={'Unnamed: 15': 'GSTR_1_5_Filing_Date'}, inplace=True)
    df3.rename(columns={'Unnamed: 16': 'GSTR_1_5_Filing_Period'}, inplace=True)
    df3.rename(columns={'Unnamed: 17': 'GSTR_3B_Filing_Status'}, inplace=True)
    df3.rename(columns={'Unnamed: 18': 'Amendment_made_if_any'}, inplace=True)
    df3.rename(columns={'Unnamed: 19': 'Tax_Period_in_which_Amended'}, inplace=True)
    df3.rename(columns={'Unnamed: 20': 'Effective_date_of_cancellation'}, inplace=True)
    df3.rename(columns={'Unnamed: 21': 'Source'}, inplace=True)
    df3.rename(columns={'Unnamed: 22': 'IRN'}, inplace=True)
    df3.rename(columns={'Unnamed: 23': 'IRN_Date'}, inplace=True)

    # here we will remove the rows, in which the invoice number has  a total
    filt = df3['Inv_CN_DN_Number_Original'].str.contains('Total', na=False)
    df3 = df3[~filt]

    df3['Inv_CN_DN_Date_Text'] = df3['Inv_CN_DN_Date_Original'].str.replace("-", ".")
    df3['Total_Tax'] = df3['IGST_Rs'] + df3['CGST_Rs'] + df3['SGST_Rs']
    df3['Unique_ID'] = df3['GSTIN_of_Supplier'] + "/" + df3['Inv_CN_DN_Number_Original'] + "/" + df3[
        'Inv_CN_DN_Date_Text']

    df3['Sheet_Name'] = ("B2B")

    df3['PAN_Number'] = df3["GSTIN_of_Supplier"].apply(lambda x: x[2:12:1])

    df3 = df3.replace(np.nan, "", regex=True)


    # A.2 : This will iterate through the B2BA file

    print("We are working on B2BA sheet of all the monthly GSTR2A..Please wait")

    df2 = pd.DataFrame()

    for files in filenames:
        df = pd.read_excel(files, sheet_name=2)

        df1 = df.drop([0, 1, 2, 3, 4, 5], axis=0)

        df1 = df1.dropna(how='all')

        df1['File_name'] = files

        df2 = df2.append(df1)

    # this is used for deleting all the rows which are totally blank
    df4 = df2.dropna(how='all')

    # this is used for renaming the names of the columns

    df4.rename(
        columns={'                                      Goods and Services Tax - GSTR-2A': 'Inv_CN_DN_Number_Original'},
        inplace=True)
    df4.rename(columns={'Unnamed: 1': 'Inv_CN_DN_Date_Original'}, inplace=True)
    df4.rename(columns={'Unnamed: 2': 'GSTIN_of_Supplier'}, inplace=True)
    df4.rename(columns={'Unnamed: 3': 'Legal_Name_Of Supplier'}, inplace=True)
    df4.rename(columns={'Unnamed: 4': 'Inv_CN_DN_Type_Revised'}, inplace=True)
    df4.rename(columns={'Unnamed: 5': 'Inv_CN_DN_Number_Revised'}, inplace=True)
    df4.rename(columns={'Unnamed: 6': 'Inv_CN_DN_Date_Revised'}, inplace=True)
    df4.rename(columns={'Unnamed: 7': 'Inv_CN_DN_Value_Revised'}, inplace=True)
    df4.rename(columns={'Unnamed: 8': 'Place_Of_Supply'}, inplace=True)
    df4.rename(columns={'Unnamed: 9': 'Supply_Attract_Reverse_Charge'}, inplace=True)
    df4.rename(columns={'Unnamed: 10': 'GST_Rate'}, inplace=True)
    df4.rename(columns={'Unnamed: 11': 'Taxable_Value_Rs'}, inplace=True)
    df4.rename(columns={'Unnamed: 12': 'IGST_Rs'}, inplace=True)
    df4.rename(columns={'Unnamed: 13': 'CGST_Rs'}, inplace=True)
    df4.rename(columns={'Unnamed: 14': 'SGST_Rs'}, inplace=True)
    df4.rename(columns={'Unnamed: 15': 'Cess'}, inplace=True)
    df4.rename(columns={'Unnamed: 16': 'GSTR_1_5_Filing_Status'}, inplace=True)
    df4.rename(columns={'Unnamed: 17': 'GSTR_1_5_Filing_Date'}, inplace=True)
    df4.rename(columns={'Unnamed: 18': 'GSTR_1_5_Filing_Period'}, inplace=True)
    df4.rename(columns={'Unnamed: 19': 'GSTR_3B_Filing_Status'}, inplace=True)
    df4.rename(columns={'Unnamed: 20': 'Effective_date_of_cancellation'}, inplace=True)
    df4.rename(columns={'Unnamed: 21': 'Amendment_made_if_any'}, inplace=True)
    df4.rename(columns={'Unnamed: 22': 'Original_tax_period_in_which_reported'}, inplace=True)

    # here we will remove the rows, in which the invoice number has  a total
    filt = df4['Inv_CN_DN_Number_Revised'].str.contains('Total', na=False)

    df4 = df4[~filt]

    df4['Inv_CN_DN_Date_Text'] = df4['Inv_CN_DN_Date_Original'].str.replace("-", ".")
    df4['Total_Tax'] = df4['IGST_Rs'] + df4['CGST_Rs'] + df4['SGST_Rs']
    df4['Unique_ID'] = df4['GSTIN_of_Supplier'] + "/" + df4['Inv_CN_DN_Number_Original'] + "/" + df4[
        'Inv_CN_DN_Date_Text']
    df4["Inv_CN_DN_Date_Revised_Unique"] = df4['Inv_CN_DN_Date_Revised'].str.replace("-", ".")

    df4['Sheet_Name'] = ("B2BA")

    df4['PAN_Number'] = df4["GSTIN_of_Supplier"].apply(lambda x: x[2:12:1])

    df4 = df4.replace(np.nan, "", regex=True)

    # A.3 : This will iterate through the CDNR file

    print("We are working on CDNR sheet of all the monthly GSTR2A..Please wait")

    df2 = pd.DataFrame()

    for files in filenames:
        df = pd.read_excel(files, sheet_name=3)

        df1 = df.drop([0, 1, 2, 3, 4], axis=0)

        df1 = df1.dropna(how='all')

        df1['File_name'] = files

        df2 = df2.append(df1)

    # this is used for deleting all the rows which are totally blank
    df5 = df2.dropna(how='all')

    # this is used for renaming the names of the columns

    df5.rename(
        columns={'                                             Goods and Services Tax - GSTR-2A': 'GSTIN_of_Supplier'},
        inplace=True)
    df5.rename(columns={'Unnamed: 1': 'Legal_Name_Of Supplier'}, inplace=True)
    df5.rename(columns={'Unnamed: 2': 'Credit_Debit_Note_Original'}, inplace=True)
    df5.rename(columns={'Unnamed: 3': 'Inv_CN_DN_Number_Original'}, inplace=True)
    df5.rename(columns={'Unnamed: 4': 'Inv_CN_DN_Type_Original'}, inplace=True)
    df5.rename(columns={'Unnamed: 5': 'Inv_CN_DN_Date_Original'}, inplace=True)
    df5.rename(columns={'Unnamed: 6': 'Inv_CN_DN_Value_Original'}, inplace=True)
    df5.rename(columns={'Unnamed: 7': 'Place_Of_Supply'}, inplace=True)
    df5.rename(columns={'Unnamed: 8': 'Supply_Attract_Reverse_Charge'}, inplace=True)
    df5.rename(columns={'Unnamed: 9': 'GST_Rate'}, inplace=True)
    df5.rename(columns={'Unnamed: 10': 'Taxable_Value_Rs'}, inplace=True)
    df5.rename(columns={'Unnamed: 11': 'IGST_Rs'}, inplace=True)
    df5.rename(columns={'Unnamed: 12': 'CGST_Rs'}, inplace=True)
    df5.rename(columns={'Unnamed: 13': 'SGST_Rs'}, inplace=True)
    df5.rename(columns={'Unnamed: 14': 'Cess'}, inplace=True)
    df5.rename(columns={'Unnamed: 15': 'GSTR_1_5_Filing_Status'}, inplace=True)
    df5.rename(columns={'Unnamed: 16': 'GSTR_1_5_Filing_Date'}, inplace=True)
    df5.rename(columns={'Unnamed: 17': 'GSTR_1_5_Filing_Period'}, inplace=True)
    df5.rename(columns={'Unnamed: 18': 'GSTR_3B_Filing_Status'}, inplace=True)
    df5.rename(columns={'Unnamed: 19': 'Amendment_made_if_any'}, inplace=True)
    df5.rename(columns={'Unnamed: 20': 'Tax_Period_in_which_Amended'}, inplace=True)
    df5.rename(columns={'Unnamed: 21': 'Effective_date_of_cancellation'}, inplace=True)
    df5.rename(columns={'Unnamed: 22': 'Source'}, inplace=True)
    df5.rename(columns={'Unnamed: 23': 'IRN'}, inplace=True)
    df5.rename(columns={'Unnamed: 24': 'IRN_Date'}, inplace=True)

    # here we will remove the rows, in which the invoice number has  a total
    filt = df5['Inv_CN_DN_Number_Original'].str.contains('Total', na=False)

    df5 = df5[~filt]

    df5['Inv_CN_DN_Date_Text'] = df5['Inv_CN_DN_Date_Original'].str.replace("-", ".")
    df5['Total_Tax'] = df5['IGST_Rs'] + df5['CGST_Rs'] + df5['SGST_Rs']
    df5['Unique_ID'] = df5['GSTIN_of_Supplier'] + "/" + df5['Inv_CN_DN_Number_Original'] + "/" + df5[
        'Inv_CN_DN_Date_Text']

    df5['Sheet_Name'] = ("CDNR")

    df5['PAN_Number'] = df5["GSTIN_of_Supplier"].apply(lambda x: x[2:12:1])

    df5 = df5.replace(np.nan, "", regex=True)

    # A.2 : This will iterate through the CDNRA file

    print("We are working on CDNRA sheet of all the monthly GSTR2A..Please wait")

    df2 = pd.DataFrame()

    for files in filenames:
        df = pd.read_excel(files, sheet_name=4)

        df1 = df.drop([0, 1, 2, 3, 4, 5], axis=0)

        df1 = df1.dropna(how='all')

        df1['File_name'] = files

        df2 = df2.append(df1)

    # this is used for deleting all the rows which are totally blank
    df6 = df2.dropna(how='all')

    # this is used for renaming the names of the columns

    df6.rename(columns={'                             Goods and Services Tax - GSTR2A': 'Credit_Debit_Note_Original'},
               inplace=True)
    df6.rename(columns={'Unnamed: 1': 'Inv_CN_DN_Number_Original'}, inplace=True)
    df6.rename(columns={'Unnamed: 2': 'Inv_CN_DN_Date_Original'}, inplace=True)
    df6.rename(columns={'Unnamed: 3': 'GSTIN_of_Supplier'}, inplace=True)
    df6.rename(columns={'Unnamed: 4': 'Legal_Name_Of Supplier'}, inplace=True)
    df6.rename(columns={'Unnamed: 5': 'Credit_Debit_Note_Revised'}, inplace=True)
    df6.rename(columns={'Unnamed: 6': 'Inv_CN_DN_Number_Revised'}, inplace=True)
    df6.rename(columns={'Unnamed: 7': 'Inv_CN_DN_Type_Revised'}, inplace=True)
    df6.rename(columns={'Unnamed: 8': 'Inv_CN_DN_Date_Revised'}, inplace=True)
    df6.rename(columns={'Unnamed: 9': 'Inv_CN_DN_Value_Revised'}, inplace=True)
    df6.rename(columns={'Unnamed: 10': 'Place_Of_Supply'}, inplace=True)
    df6.rename(columns={'Unnamed: 11': 'Supply_Attract_Reverse_Charge'}, inplace=True)
    df6.rename(columns={'Unnamed: 12': 'GST_Rate'}, inplace=True)
    df6.rename(columns={'Unnamed: 13': 'Taxable_Value_Rs'}, inplace=True)
    df6.rename(columns={'Unnamed: 14': 'IGST_Rs'}, inplace=True)
    df6.rename(columns={'Unnamed: 15': 'CGST_Rs'}, inplace=True)
    df6.rename(columns={'Unnamed: 16': 'SGST_Rs'}, inplace=True)
    df6.rename(columns={'Unnamed: 17': 'Cess'}, inplace=True)
    df6.rename(columns={'Unnamed: 18': 'GSTR_1_5_Filing_Status'}, inplace=True)
    df6.rename(columns={'Unnamed: 19': 'GSTR_1_5_Filing_Date'}, inplace=True)
    df6.rename(columns={'Unnamed: 20': 'GSTR_1_5_Filing_Period'}, inplace=True)
    df6.rename(columns={'Unnamed: 21': 'GSTR_3B_Filing_Status'}, inplace=True)
    df6.rename(columns={'Unnamed: 22': 'Amendment_made_if_any'}, inplace=True)
    df6.rename(columns={'Unnamed: 23': 'Original_tax_period_in_which_reported'}, inplace=True)
    df6.rename(columns={'Unnamed: 24': 'Effective_date_of_cancellation'}, inplace=True)

    # here we will remove the rows, in which the invoice number has  a total
    filt = df6['Inv_CN_DN_Number_Revised'].str.contains('Total', na=False)

    df6 = df6[~filt]

    df6['Inv_CN_DN_Date_Text'] = df6['Inv_CN_DN_Date_Original'].str.replace("-", ".")
    df6['Total_Tax'] = df6['IGST_Rs'] + df6['CGST_Rs'] + df6['SGST_Rs']
    df6['Unique_ID'] = df6['GSTIN_of_Supplier'] + "/" + df6['Inv_CN_DN_Number_Original'] + "/" + df6[
        'Inv_CN_DN_Date_Text']

    df6["Inv_CN_DN_Date_Revised_Unique"] = df6['Inv_CN_DN_Date_Revised'].str.replace("-", ".")

    df6['Sheet_Name'] = ("CDNRA")

    df6['PAN_Number'] = df6["GSTIN_of_Supplier"].apply(lambda x: x[2:12:1])

    df6 = df6.replace(np.nan, "", regex=True)

    # Making a combined sheet with all merged

    print("We are Combining B2B, B2BA, CDNR & CDNRA in 1 sheet...Please wait..!")

    df8 = df3.append(df4)

    df9 = df8.append(df5)

    df10 = df9.append(df6)

    df10['PAN_Number'] = df10["GSTIN_of_Supplier"].apply(lambda x: x[2:12:1])

    df10 = df10.replace(np.nan, "", regex=True)

    df10["Ultimate_Unique"] = df10["Sheet_Name"] + "/" + df10["Supply_Attract_Reverse_Charge"] + df10[
        "GSTR_1_5_Filing_Status"] + "/" + df10["Unique_ID"]

    df10["PAN_3_Way_Key"] = np.where(df10["Sheet_Name"] == "B2BA",
                                     df10["PAN_Number"] + "/" + df10["Inv_CN_DN_Number_Revised"] + "/"
                                     + df10["Inv_CN_DN_Date_Revised_Unique"],
                                     df10["PAN_Number"] + "/" + df10["Inv_CN_DN_Number_Original"]
                                     + "/" + df10["Inv_CN_DN_Date_Text"])

    df10["PAN_2_Way_Key_PAN_InvNo"] = np.where(df10["Sheet_Name"] == "B2BA",
                                               df10["PAN_Number"] + "/" + df10["Inv_CN_DN_Number_Revised"]
                                               , df10["PAN_Number"] + "/" + df10["Inv_CN_DN_Number_Original"])

    df10["PAN_2_Way_Key_PAN_InvDt"] = np.where(df10["Sheet_Name"] == "B2BA",
                                               df10["PAN_Number"] + "/" + df10["Inv_CN_DN_Date_Revised_Unique"]
                                               , df10["PAN_Number"] + "/" + df10["Inv_CN_DN_Date_Text"])

    # maiking a sheet with person who did not file the GSTR 1

    df11 = df10[df10['GSTR_1_5_Filing_Status'] == "N"]

    df12 = df10[(df10['Supply_Attract_Reverse_Charge'] == "Y") & (df10['GSTR_1_5_Filing_Status'] == "Y")]

    df13 = df10[(df10['Supply_Attract_Reverse_Charge'] == "N") & (df10['GSTR_1_5_Filing_Status'] == "Y") & (
            df10['Total_Tax'] < 1)]

    df14 = df10[(df10['Supply_Attract_Reverse_Charge'] == "N") & (df10['GSTR_1_5_Filing_Status'] == "Y") & (
            df10['Total_Tax'] >= 1)]

    # saving the file with the name "Combined"

    extension = os.path.splitext(filepath)[1]
    filename = os.path.splitext(filepath)[0]
    pth = os.path.dirname(filepath)
    newfile = os.path.join(pth, "Merged" + 'GSTR2A_all_Files_by_Effcorp' + extension)

    writer = pd.ExcelWriter(newfile, engine='openpyxl')

    print("Please wait.. we are creating different sheets and finalizing the file.....")

    df3.to_excel(writer, sheet_name="B2B")

    df4.to_excel(writer, sheet_name="B2BA")

    df5.to_excel(writer, sheet_name="CDNR")

    df6.to_excel(writer, sheet_name="CDNRA")

    titles = list(df10.columns)

    titles[0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], \
    titles[10], titles[11], titles[12], titles[13], titles[14], titles[15], titles[16], titles[17], titles[18], titles[
        19], titles[20], titles[21], titles[22], titles[23], titles[24], titles[25], titles[26], titles[27], titles[28], \
    titles[29], titles[30], titles[31], titles[32], titles[33], titles[34], titles[35] = titles[24], titles[28], titles[
        0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], titles[
                                                                                             10], titles[11], titles[
                                                                                             12], titles[13], titles[
                                                                                             26], titles[25], titles[
                                                                                             21], titles[27], titles[
                                                                                             14], titles[15], titles[
                                                                                             16], titles[17], titles[
                                                                                             18], titles[19], titles[
                                                                                             20], titles[22], titles[
                                                                                             23], titles[29], titles[
                                                                                             30], titles[31], titles[
                                                                                             32], titles[33], titles[
                                                                                             34], titles[35]

    df10[titles].to_excel(writer, sheet_name="All_Combined")

    titles = list(df11.columns)

    titles[0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], \
    titles[10], titles[11], titles[12], titles[13], titles[14], titles[15], titles[16], titles[17], titles[18], titles[
        19], titles[20], titles[21], titles[22], titles[23], titles[24], titles[25], titles[26], titles[27], titles[28], \
    titles[29], titles[30], titles[31], titles[32], titles[33], titles[34], titles[35] = titles[24], titles[28], titles[
        0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], titles[
                                                                                             10], titles[11], titles[
                                                                                             12], titles[13], titles[
                                                                                             26], titles[25], titles[
                                                                                             21], titles[27], titles[
                                                                                             14], titles[15], titles[
                                                                                             16], titles[17], titles[
                                                                                             18], titles[19], titles[
                                                                                             20], titles[22], titles[
                                                                                             23], titles[29], titles[
                                                                                             30], titles[31], titles[
                                                                                             32], titles[33], titles[
                                                                                             34], titles[35]

    df11[titles].to_excel(writer, sheet_name="GSTR_1_Not Filed")

    titles = list(df12.columns)

    titles[0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], \
    titles[10], titles[11], titles[12], titles[13], titles[14], titles[15], titles[16], titles[17], titles[18], titles[
        19], titles[20], titles[21], titles[22], titles[23], titles[24], titles[25], titles[26], titles[27], titles[28], \
    titles[29], titles[30], titles[31], titles[32], titles[33], titles[34], titles[35] = titles[24], titles[28], titles[
        0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], titles[
                                                                                             10], titles[11], titles[
                                                                                             12], titles[13], titles[
                                                                                             26], titles[25], titles[
                                                                                             21], titles[27], titles[
                                                                                             14], titles[15], titles[
                                                                                             16], titles[17], titles[
                                                                                             18], titles[19], titles[
                                                                                             20], titles[22], titles[
                                                                                             23], titles[29], titles[
                                                                                             30], titles[31], titles[
                                                                                             32], titles[33], titles[
                                                                                             34], titles[35]

    df12[titles].to_excel(writer, sheet_name="GSTR_Filed_RCM_Yes")

    titles = list(df13.columns)

    titles[0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], \
    titles[10], titles[11], titles[12], titles[13], titles[14], titles[15], titles[16], titles[17], titles[18], titles[
        19], titles[20], titles[21], titles[22], titles[23], titles[24], titles[25], titles[26], titles[27], titles[28], \
    titles[29], titles[30], titles[31], titles[32], titles[33], titles[34], titles[35] = titles[24], titles[28], titles[
        0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], titles[
                                                                                             10], titles[11], titles[
                                                                                             12], titles[13], titles[
                                                                                             26], titles[25], titles[
                                                                                             21], titles[27], titles[
                                                                                             14], titles[15], titles[
                                                                                             16], titles[17], titles[
                                                                                             18], titles[19], titles[
                                                                                             20], titles[22], titles[
                                                                                             23], titles[29], titles[
                                                                                             30], titles[31], titles[
                                                                                             32], titles[33], titles[
                                                                                             34], titles[35]

    df13[titles].to_excel(writer, sheet_name="Tax_Zero_Cases")

    titles = list(df14.columns)

    titles[0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], \
    titles[10], titles[11], titles[12], titles[13], titles[14], titles[15], titles[16], titles[17], titles[18], titles[
        19], titles[20], titles[21], titles[22], titles[23], titles[24], titles[25], titles[26], titles[27], titles[28], \
    titles[29], titles[30], titles[31], titles[32], titles[33], titles[34], titles[35] = titles[24], titles[28], titles[
        0], titles[1], titles[2], titles[3], titles[4], titles[5], titles[6], titles[7], titles[8], titles[9], titles[
                                                                                             10], titles[11], titles[
                                                                                             12], titles[13], titles[
                                                                                             26], titles[25], titles[
                                                                                             21], titles[27], titles[
                                                                                             14], titles[15], titles[
                                                                                             16], titles[17], titles[
                                                                                             18], titles[19], titles[
                                                                                             20], titles[22], titles[
                                                                                             23], titles[29], titles[
                                                                                             30], titles[31], titles[
                                                                                             32], titles[33], titles[
                                                                                             34], titles[35]

    df14[titles].to_excel(writer, sheet_name="Working_Cases")

    writer.save()
    writer.close()

    print(f"All files have been combined and a single file named {newfile} has been created")

    return (writer)





def download(pth=os.getcwd()):



    """
    This is a function to download the GSTR2A & ITR format and also instructions to use this utility


    :param pth: This takes a single argument which is a pathn which the user wants to store the Format files

    :param Type: This parameter is a optional argument

                In case the parameter is not provided, the current working directory is taken as the Pth and the fomrat os downloaded in that folder


    :return writer: This function will return a excel file which has a format for the reconciliation of the GSTR2A and the ITR

    This is a dependednt function for the next main function reco_itr_2a. 

    There are mandatory columns and it has to be ensured that the names of the Mandatory columns are same as in the format

    There is no requirment for the sequence of the columns to be same as the Format

    The excel file in which the data is kept can hae multiple sheets , but the nme of the sheet should be same as in the format

    For more details, refer the Sheet "Important_Checklist" downloaded in the format


    """



    import pandas as pd
    import numpy as np
    import openpyxl


    fullpath1 = pth + "\\" + "Formats.xlsx"
    print(f"The path selected is {fullpath1}")

    writer = pd.ExcelWriter(fullpath1, engine='xlsxwriter', options={'strings_to_formulas': True})

    dict1 = {"Vendor_GST_REG": ["Mandatory"], "Vendor_Name": ["Optional"], "Invoice_Number": ["Mandatory"],
             "Invoice_Date_Text": ["Mandatory"], "Total_Tax": ["Mandatory"], "IGST": ["Optional"], "SGST": ["Optional"],
             "CGST": ["Optional"], "UTGST": ["Optional"],"User Defined1":["Optional"],"User Defined2":["Optional"],"User Defined3":["Optional"],"User Defined4":["Optional"]}

    df1 = pd.DataFrame(dict1)
    df1.to_excel(writer, sheet_name="Main_ITR_Format", index=False)

    dict2={"GSTIN_of_Supplier":["Mandatory"],"Inv_CN_DN_Number_Final":["Mandatory"],"Legal_Name_Of Supplier":["Optional"],"Inv_CN_DN_Date_Text":["Mandatory"],"Total_Tax":["Mandatory"],"GSTR_1_5_Filing_Status":["Mandatory"],"Supply_Attract_Reverse_Charge":["Mandatory"],"IGST":["Optional"],"SGST":["Optional"],"CGST":["Optional"],"UTGST":["Optional"],"User Defined1":["Optional"],"User Defined2":["Optional"],"User Defined3":["Optional"],"User Defined4":["Optional"]}
    df2 = pd.DataFrame(dict2)
    df2.to_excel(writer, sheet_name="Main_2A_Format", index=False)

    dict3={"Things to ensure before Running the Program":[" ","Keep GSTR2A and ITR in different Folders ","In that Folder , there should not be any other excel files",
    "Format of the ITR & GSTR2A can be as per the user ","However, below points to be taken care","The name of the Sheet having the  ITR should be Main_ITR_Format ",
    "The name of the Sheet having the Consolidated GSTR2A should be Main_2A_Format "," ","There are 6 Mandatory columns in  GSTR2A and 4 Mandatory columns in ITR ","The Name to be assigned to these 6 Mandatory columns must be same as in the format",
    "In ITR , Mandatory columns are Vendor_GST_REG , Invoice_Number, Invoice_Date_Text, Total_Tax","Even the upper and Lower case should be same as in the Format",
    "In GSTR2A , Mandatory columns are GSTIN_of_Supplier , Inv_CN_DN_Number_Final, Inv_CN_DN_Date_Text, Total_Tax,Supply_Attract_Reverse_Charge,GSTR_1_5_Filing_Status","Take care of the Upper and Lower case and special Character",
    "The sequence of the Columns is not relevant. User can maintain the Sequence of the columns as per his own convinience","The ITR or GSTR2A file can also have multiple other sheets as per need of user , but relevant data for matching should be in one sheet only .",
    "But it has to be ensured that the main sheet in the GSTR2A and ITR has exactly the same namee as mentioned in the Format", "  "," ","For any issues in running the Code, send your issues to efficientcorporates.info@gmail.com",
    "For more such Automation Videos, Follow YouTube Channel Efficient Corporates"]}
    
    df3 = pd.DataFrame(dict3)
    df3.to_excel(writer, sheet_name="Important_Checklist", index=False)


    writer.save()

    writer.close()

    print(f'The Formats have been saved in below path \n {fullpath1}\n ')

    return (writer)






def reco_itr_2a(files_itr,files_con2a,tol_limit=100):

    """
    This fucntion is for reconciling the GSTr2A and the ITR

    Please download the Format using the Function download and go through the Important Checklist

    This function takes the 3 parameters. Two are Mandatory and 1 is optional

    :param files_itr: This argument should be the complete path to the ITR file which is as per the format 

                    Please ensure to provide the compelte filepath of ITR till the extension

    :param files_con2a : This is the argument for the complete filepath of the GSTR2A file.
                        
                        Please ensure to gve the complete file path till the extension

    :param tol_limit : This is also next important parameter. This is the Tolerance limit.

                        If a invoice is booked with Tax of Rs 12,300 , but the same invoice is given in GSTR2A as Rs 12450.

                        Now, there is a difference of Rs 150. Now , if the tolerance limit is kept as 100, then this case will be considered NOT MATCHING

                        But, if the tolerance limit is kept as 200, then this case will be considered as a match

                        Use can provide the Tolerance limit value based on the size of the client


                        If no parameter is provided , then the 100 is taken as the Tolerance limit

    :return output : This function will return 2 files 1) Summary.xlsx and 2) Working.xlsx

                    These 2 files will be stored in the folder where the Combined GSTR2A is stored

                    The Summary fil will ave a snapshot of the matching exercise and will tell the Total cases, matched cases and the unmatched cases

                    The matching is done under 7 different categories

                    a) GST+INV NO + INV Date +Tax Amount >> Complete 3 way match

                    b)GST + INV NO +Tax Amount >> Complete 2 way match

                    c)GST + INV Date +Tax Amount >> Complete 2 way match

                    d)PAN+ INV NO + INV Date +Tax Amount >> Complete 3 way match

                    e)PAN+INV NO  +Tax Amount >> Complete 2 way match

                    f) PAN + INV Date +Tax Amount >> Complete 2 way match

                    g) Fuzzy Look up Match: These are the cases with 90% Invoice Number and 100% PAN Number matching . Just the Invoice Number matches , not the Tax Amount or date


                    Also, the Unmatched cases of ITR will be bifurcated into 3 difefrent buckets 

                    1. Cases whose GST/PAN is not present in GSTR2A (No Scope of Mathing)
                    2. Cases where the GST Number entered in Purchase Register is INVALID
                    3. Cases where the Invoice Number is Invalid
                    4. Other Remaining Unmatched Cases 



    """



    import numpy as np
    import openpyxl

    warnings.filterwarnings('ignore')



    print(f'The Consolidated GSTR2A file path is {files_con2a}')
    print(f'The ITR file path is {files_itr}')

    pth = os.path.dirname(str(files_con2a))

    fullpath1 = pth + "/" + "Workings.xlsx"

    writer = pd.ExcelWriter(fullpath1, engine='xlsxwriter', options={'strings_to_formulas': True})
    #




    fullpath1a = pth + "/" + "Summary.xlsx"
    writer1 = pd.ExcelWriter(fullpath1a, engine='xlsxwriter', options={'strings_to_formulas': True})

    df1 = pd.DataFrame()
    df1.to_excel(writer1, sheet_name="Summary", index=False)

    writer1.save()

    fullpath2 = fullpath1a.replace("/", "\\")  # this is a very useful command for defining the correct filepath

    wb = load_workbook(fullpath2)
    ws = wb["Summary"]

    ws["B2"].value = "SUMMARY OF THE RECONCILIATION OF GSTR2A Vs ITR"
    ws.merge_cells("B2:F2")
    ws["C4"].value = "GSTR2A"
    ws.merge_cells("C4:D4")
    ws["E4"].value = "Purchase Register"
    ws.merge_cells("E4:F4")

    ws["B4"].value = "Particulars"
    ws.merge_cells("B4:B5")
    ws["C5"].value = "Count"
    ws["D5"].value = "Tax Amount"
    ws["E5"].value = "Count"
    ws["F5"].value = "Tax Amount"

    ws["B7"].value = "Total cases in Original Files"
    ws["B8"].value = "Less: No GST Number in Purchase Register"
    ws["B9"].value = "Less: GSTR-1 Not filed cases"
    ws["B10"].value = "Less: GSTR-1 Filed but RCM cases "
    ws["B11"].value = "Less: No Invoice Number in Purchase Register"



    ws["B12"].value = "Net cases to be Matched"
    ws["B14"].value = "Matched with GST_INVNO_INVDATE_3_WAY"
    ws["B15"].value = "Matched with GST_INVNO_2_WAY"
    ws["B16"].value = "Matched with GST_INVDATE_2_WAY"

    ws["B18"].value = "Identified Possible Matches - Fuzzy Logic"

    ws["B20"].value = "Matched with PAN_INVNO_INVDATE_3_WAY"
    ws["B21"].value = "Matched with PAN_INVNO_2_WAY"
    ws["B22"].value = "Matched with PAN_INVDATE_2_WAY"

    ws["B24"].value = "Unmatched Cases"


    ws["B25"].value = "Unmatched Cases -PAN/GST not available in GSTR2A"
    ws["B26"].value = "Unmatched Cases with Invalid GSTIN"


    ws["B27"].value = "Unmatched cases with Invalid Invoice Number"

    ws["B28"].value = "Other Unmatched Cases"

    ws["B30"].value = "Check"

    # setting the tolerance limit for matching in Rupees

    tol_limit = int(tol_limit)

    print(f"The tolerance limit is set to {tol_limit}")

    ws["F1"].value = f"Tolerance Limit was {tol_limit}"

    gstr2a = pd.read_excel(files_con2a, sheet_name="Main_2A_Format",dtype={"Inv_CN_DN_Number_Final":str, "Inv_CN_DN_Date_Text":str, "Total_Tax":float})

    try:
        gstr2a['Inv_CN_DN_Number_Finall'] = gstr2a["Inv_CN_DN_Number_Final"].apply(lambda x: x.lower(str()))
    except:
        gstr2a['Inv_CN_DN_Number_Finall'] = gstr2a["Inv_CN_DN_Number_Final"]

    gstr2a['GST_INVNO_INVDATE_3_WAY'] = gstr2a['GSTIN_of_Supplier'] + "/" + gstr2a['Inv_CN_DN_Number_Finall'] + "/" + gstr2a['Inv_CN_DN_Date_Text']

    gstr2a['GST_INVNO_2_WAY'] = gstr2a['GSTIN_of_Supplier'] + "/" + gstr2a['Inv_CN_DN_Number_Finall']

    gstr2a['GST_INVDATE_2_WAY'] = gstr2a['GSTIN_of_Supplier'] + "/" + gstr2a['Inv_CN_DN_Date_Text']

    
    gstr2a['PAN_Number'] = gstr2a["GSTIN_of_Supplier"].apply(lambda x:extract_pan(x))
    
    # the PAN number matches will be used as possible matches

    gstr2a['PAN_INVNO_INVDATE_3_WAY'] = gstr2a['PAN_Number'] + "/" + gstr2a['Inv_CN_DN_Number_Finall'] + "/" + \
                                        gstr2a['Inv_CN_DN_Date_Text']

    gstr2a['PAN_INVNO_2_WAY'] = gstr2a['PAN_Number'] + "/" + gstr2a['Inv_CN_DN_Number_Finall']

    gstr2a['PAN_INVDATE_2_WAY'] = gstr2a['PAN_Number'] + "/" + gstr2a['Inv_CN_DN_Date_Text']


    itr = pd.read_excel(files_itr, sheet_name="Main_ITR_Format",dtype={"Invoice_Number":str, "Invoice_Date_Text":str,"Total_Tax":float})

    try:
        itr["Invoice_Numberl"] = itr["Invoice_Number"].apply(lambda x: x.lower(str()))
    except:
        itr["Invoice_Numberl"] = itr["Invoice_Number"]

    itr["GST_INVNO_INVDATE_3_WAY"] = itr["Vendor_GST_REG"] + "/" + itr["Invoice_Numberl"] + "/" + itr[
        "Invoice_Date_Text"]

    itr["GST_INVNO_2_WAY"] = itr["Vendor_GST_REG"] + "/" + itr["Invoice_Numberl"]

    itr["GST_INVDATE_2_WAY"] = itr["Vendor_GST_REG"] + "/" + itr["Invoice_Date_Text"]

   
    itr["PAN_Number"] = itr["Vendor_GST_REG"].apply(lambda x:extract_pan(x))
    


    # the PAN number matches will be used as possible matches

    itr["PAN_INVNO_INVDATE_3_WAY"] = itr["PAN_Number"] + "/" + itr["Invoice_Numberl"] + "/" + itr["Invoice_Date_Text"]

    itr["PAN_INVNO_2_WAY"] = itr["PAN_Number"] + "/" + itr["Invoice_Numberl"]

    itr["PAN_INVDATE_2_WAY"] = itr["PAN_Number"] + "/" + itr["Invoice_Date_Text"]

    ws["C7"].value = list(gstr2a.shape)[0]
    ws["D7"].value = sum(gstr2a["Total_Tax"])
    ws["E7"].value = list(itr.shape)[0]
    ws["F7"].value = sum(itr["Total_Tax"])



    #data Cleaning for GSTr2A:

    try:
        
        gstr2a_not_filed=gstr2a[gstr2a['GSTR_1_5_Filing_Status'] == "N"]

    except:

        pass


    try:
        gstr2a_rcm=gstr2a[(gstr2a['Supply_Attract_Reverse_Charge'] == "Y") & (gstr2a['GSTR_1_5_Filing_Status'] == "Y")]
    except:
        pass


    try:
        gstr2a_work=gstr2a[(gstr2a['Supply_Attract_Reverse_Charge'] != "Y") & (gstr2a['GSTR_1_5_Filing_Status'] == "Y")]
    except:
        gstr2a_work=gstr_2a
        

    ws["C9"].value = len(gstr2a_not_filed["GSTIN_of_Supplier"])
    ws["D9"].value = sum(gstr2a_not_filed["Total_Tax"])
    ws["C10"].value = len(gstr2a_rcm["GSTIN_of_Supplier"])
    ws["D10"].value = sum(gstr2a_rcm["Total_Tax"])

    ws["C12"].value = len(gstr2a_work["GSTIN_of_Supplier"])
    ws["D12"].value = sum(gstr2a_work["Total_Tax"])


    #data cleaing  for ITR as of now is only Blank GST Reg No and Blank Invoice Number.
    #So, net case to be matched will equal to Total cases

    mask=itr["Vendor_GST_REG"].isnull()

    try:
        itr_nogst=itr[mask]
    except:
        pass


    itr_gst=itr[~mask]

    mask2=itr_gst["Invoice_Numberl"].isnull()

    try:
        itr_noinvno=itr_gst[mask2]
    except:
        pass

    itr_work=itr_gst[~mask2]


    ws["E8"].value = list(itr_nogst.shape)[0]
    ws["F8"].value = sum(itr_nogst["Total_Tax"])

    ws["E11"].value = list(itr_noinvno.shape)[0]
    ws["F11"].value = sum(itr_noinvno["Total_Tax"])




    ws["E12"].value = list(itr_work.shape)[0]
    ws["F12"].value = sum(itr_work["Total_Tax"])




    # First Cut Matching : Here we will try to do that Matching based on 3 way i.e GST No, Inv No & Inv Date being same in ITR & GSTR2A



    gstr2a_pivot = pd.pivot_table(gstr2a_work, values="Total_Tax", index=["GST_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

    itr_pivot = pd.pivot_table(itr_work, values="Total_Tax", index=["GST_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="GST_INVNO_INVDATE_3_WAY", right_on="GST_INVNO_INVDATE_3_WAY",
                                  how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks_Effcorp"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks_Effcorp"].values == "Exact Match within Tolerance"

    match_3_way_list = compared[mask_1]["GST_INVNO_INVDATE_3_WAY"].values

    mask_1a = gstr2a["GST_INVNO_INVDATE_3_WAY"].isin(match_3_way_list)  # returns a Boolean Array

    mask_1b = itr["GST_INVNO_INVDATE_3_WAY"].isin(match_3_way_list)  # returns a boolean array

    matched_gstr2a_3way = gstr2a[mask_1a]
    matched_gstr2a_3way["Matching Category"] = "3 Way matching GST + Inv No+ Inv Date"
    matched_itr_3way = itr[mask_1b]
    matched_itr_3way["Matching Category"] = "3 Way matching GST + Inv No+ Inv Date"

    ws["C14"].value = len(matched_gstr2a_3way["GST_INVNO_INVDATE_3_WAY"])
    ws["D14"].value = sum(matched_gstr2a_3way["Total_Tax"])
    ws["E14"].value = len(matched_itr_3way["GST_INVNO_INVDATE_3_WAY"])
    ws["F14"].value = sum(matched_itr_3way["Total_Tax"])

    bal_gstr2a_1cut = gstr2a_work[~mask_1a]
    bal_itr_1cut = itr_work[~mask_1b]

    # Second Cut Matching : Here we will try to do that Matching based on 2 way i.e GST No & Inv No

    gstr2a_pivot = pd.pivot_table(bal_gstr2a_1cut, values="Total_Tax", index=["GST_INVNO_2_WAY"], aggfunc=np.sum)

    itr_pivot = pd.pivot_table(bal_itr_1cut, values="Total_Tax", index=["GST_INVNO_2_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="GST_INVNO_2_WAY", right_on="GST_INVNO_2_WAY", how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks_Effcorp"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks_Effcorp"].values == "Exact Match within Tolerance"

    match_2_way_list1 = compared[mask_1]["GST_INVNO_2_WAY"].values

    mask_1a = bal_gstr2a_1cut["GST_INVNO_2_WAY"].isin(match_2_way_list1)  # returns a Boolean Array

    mask_1b = bal_itr_1cut["GST_INVNO_2_WAY"].isin(match_2_way_list1)  # returns a boolean array

    matched_gstr2a_2way1 = bal_gstr2a_1cut[mask_1a]
    matched_itr_2way1 = bal_itr_1cut[mask_1b]

    matched_gstr2a_2way1["Matching Category"] = "2 Way matching GST + Inv No"
    matched_itr_2way1["Matching Category"] = "2 Way matching GST + Inv No"

    ws["C15"].value = len(matched_gstr2a_2way1["GST_INVNO_2_WAY"])
    ws["D15"].value = sum(matched_gstr2a_2way1["Total_Tax"])
    ws["E15"].value = len(matched_itr_2way1["GST_INVNO_2_WAY"])
    ws["F15"].value = sum(matched_itr_2way1["Total_Tax"])

    bal_gstr2a_2cut = bal_gstr2a_1cut[~mask_1a]
    bal_itr_2cut = bal_itr_1cut[~mask_1b]

    # Third Cut Matching : Here we will try to do that Matching based on 2 way i.e GST No & Inv Date

    gstr2a_pivot = pd.pivot_table(bal_gstr2a_2cut, values="Total_Tax", index=["GST_INVDATE_2_WAY"], aggfunc=np.sum)

    itr_pivot = pd.pivot_table(bal_itr_2cut, values="Total_Tax", index=["GST_INVDATE_2_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="GST_INVDATE_2_WAY", right_on="GST_INVDATE_2_WAY", how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks_Effcorp"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks_Effcorp"].values == "Exact Match within Tolerance"

    match_2_way_list2 = compared[mask_1]["GST_INVDATE_2_WAY"].values

    mask_1a = bal_gstr2a_2cut["GST_INVDATE_2_WAY"].isin(match_2_way_list2)  # returns a Boolean Array

    mask_1b = bal_itr_2cut["GST_INVDATE_2_WAY"].isin(match_2_way_list2)  # returns a boolean array

    matched_gstr2a_2way2 = bal_gstr2a_2cut[mask_1a]
    matched_itr_2way2 = bal_itr_2cut[mask_1b]

    matched_gstr2a_2way2["Matching Category"] = "2 Way matching GST + Inv Date"
    matched_itr_2way2["Matching Category"] = "2 Way matching GST + Inv Date"

    ws["C16"].value = len(matched_gstr2a_2way2["GST_INVDATE_2_WAY"])
    ws["D16"].value = sum(matched_gstr2a_2way2["Total_Tax"])
    ws["E16"].value = len(matched_itr_2way2["GST_INVDATE_2_WAY"])
    ws["F16"].value = sum(matched_itr_2way2["Total_Tax"])

    bal_gstr2a_3cut = bal_gstr2a_2cut[~mask_1a]
    bal_itr_3cut = bal_itr_2cut[~mask_1b]

    print(f"The 3 way matching using GST is done.... Now ,we are doing the mtching using PAN..Please wait...!!")




    #after the 3 cut matching, now we try to find out the Possible matches in form of PAN matching and upper /lower case matching
    # Fourth Cut Matching : Here we will try to do that Matching based on 3 way With PAN No Inv NO and Inv Date

    gstr2a_pivot = pd.pivot_table(bal_gstr2a_3cut, values="Total_Tax", index=["PAN_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

    itr_pivot = pd.pivot_table(bal_itr_3cut, values="Total_Tax", index=["PAN_INVNO_INVDATE_3_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="PAN_INVNO_INVDATE_3_WAY", right_on="PAN_INVNO_INVDATE_3_WAY", how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks_Effcorp"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks_Effcorp"].values == "Exact Match within Tolerance"

    match_3_way_list2 = compared[mask_1]["PAN_INVNO_INVDATE_3_WAY"].values

    mask_1a = bal_gstr2a_3cut["PAN_INVNO_INVDATE_3_WAY"].isin(match_3_way_list2)  # returns a Boolean Array

    mask_1b = bal_itr_3cut["PAN_INVNO_INVDATE_3_WAY"].isin(match_3_way_list2)  # returns a boolean array

    matched_gstr2a_3way2 = bal_gstr2a_3cut[mask_1a]
    matched_itr_3way2 = bal_itr_3cut[mask_1b]

    matched_gstr2a_3way2["Matching Category"] = "3 Way matching PAN + Inv No+ Inv Date"
    matched_itr_3way2["Matching Category"] = "3 Way matching PAN + Inv No + Inv Date"

    ws["C20"].value = len(matched_gstr2a_3way2["PAN_INVNO_INVDATE_3_WAY"])
    ws["D20"].value = sum(matched_gstr2a_3way2["Total_Tax"])
    ws["E20"].value = len(matched_itr_3way2["PAN_INVNO_INVDATE_3_WAY"])
    ws["F20"].value = sum(matched_itr_3way2["Total_Tax"])

    bal_gstr2a_4cut = bal_gstr2a_3cut[~mask_1a]
    bal_itr_4cut = bal_itr_3cut[~mask_1b]

    # Fifth Cut Matching : Here we will try to do that Matching based on 3 way With PAN No Inv NO and Inv Date

    gstr2a_pivot = pd.pivot_table(bal_gstr2a_4cut, values="Total_Tax", index=["PAN_INVNO_2_WAY"],
                                  aggfunc=np.sum)

    itr_pivot = pd.pivot_table(bal_itr_4cut, values="Total_Tax", index=["PAN_INVNO_2_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="PAN_INVNO_2_WAY", right_on="PAN_INVNO_2_WAY",
                                  how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks_Effcorp"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks_Effcorp"].values == "Exact Match within Tolerance"

    match_2_way_list3 = compared[mask_1]["PAN_INVNO_2_WAY"].values

    mask_1a = bal_gstr2a_4cut["PAN_INVNO_2_WAY"].isin(match_2_way_list3)  # returns a Boolean Array

    mask_1b = bal_itr_4cut["PAN_INVNO_2_WAY"].isin(match_2_way_list3)  # returns a boolean array

    matched_gstr2a_2way3 = bal_gstr2a_4cut[mask_1a]
    matched_itr_2way3 = bal_itr_4cut[mask_1b]

    matched_gstr2a_2way3["Matching Category"] = "2 Way matching PAN + Inv No"
    matched_itr_2way3["Matching Category"] = "2 Way matching PAN + Inv No "

    ws["C21"].value = len(matched_gstr2a_2way3["PAN_INVNO_2_WAY"])
    ws["D21"].value = sum(matched_gstr2a_2way3["Total_Tax"])
    ws["E21"].value = len(matched_itr_2way3["PAN_INVNO_2_WAY"])
    ws["F21"].value = sum(matched_itr_2way3["Total_Tax"])

    bal_gstr2a_5cut = bal_gstr2a_4cut[~mask_1a]
    bal_itr_5cut = bal_itr_4cut[~mask_1b]



    # Sixth Cut Matching : Here we will try to do that Matching based on 3 way With PAN No Inv NO and Inv Date

    gstr2a_pivot = pd.pivot_table(bal_gstr2a_5cut, values="Total_Tax", index=["PAN_INVDATE_2_WAY"],
                                  aggfunc=np.sum)

    itr_pivot = pd.pivot_table(bal_itr_4cut, values="Total_Tax", index=["PAN_INVDATE_2_WAY"], aggfunc=np.sum)

    gstr2a_pivot.rename(columns={'Total_Tax': 'Tax_as_per_GSTR2A'}, inplace=True)

    itr_pivot.rename(columns={'Total_Tax': 'Tax_as_per_ITR'}, inplace=True)

    gstr2a_pivot.reset_index(inplace=True)

    itr_pivot.reset_index(inplace=True)

    compared = gstr2a_pivot.merge(itr_pivot, left_on="PAN_INVDATE_2_WAY", right_on="PAN_INVDATE_2_WAY",
                                  how="left")

    compared = compared.replace(np.nan, 0, regex=True)

    compared["Difference_in_Tax"] = compared["Tax_as_per_GSTR2A"] - compared["Tax_as_per_ITR"]

    conditions = [compared["Difference_in_Tax"] > (tol_limit),

                  compared["Difference_in_Tax"] < (tol_limit * -1),

                  ((compared["Difference_in_Tax"] > (tol_limit * -1)) & (compared["Difference_in_Tax"] < (tol_limit)))

                  ]

    results = ["Excess in GSTR 2A, Less in ITR",

               "Excess in ITR, Less in GSTR2A",

               "Exact Match within Tolerance"]

    compared["Remarks_Effcorp"] = np.select(conditions, results)

    # The Above block of code gives us the pivot table with a comparison of the GSTR2A and the ITR with remarks column

    # now we will select the exact match within the Tolerance level

    mask_1 = compared["Remarks_Effcorp"].values == "Exact Match within Tolerance"

    match_2_way_list4 = compared[mask_1]["PAN_INVDATE_2_WAY"].values

    mask_1a = bal_gstr2a_5cut["PAN_INVDATE_2_WAY"].isin(match_2_way_list4)  # returns a Boolean Array

    mask_1b = bal_itr_5cut["PAN_INVDATE_2_WAY"].isin(match_2_way_list4)  # returns a boolean array

    matched_gstr2a_2way4 = bal_gstr2a_5cut[mask_1a]
    matched_itr_2way4 = bal_itr_5cut[mask_1b]

    matched_gstr2a_2way4["Matching Category"] = "2 Way matching PAN + Inv Date"
    matched_itr_2way4["Matching Category"] = "2 Way matching PAN + Inv Date "

    ws["C22"].value = len(matched_gstr2a_2way4["PAN_INVDATE_2_WAY"])
    ws["D22"].value = sum(matched_gstr2a_2way4["Total_Tax"])
    ws["E22"].value = len(matched_itr_2way4["PAN_INVDATE_2_WAY"])
    ws["F22"].value = sum(matched_itr_2way4["Total_Tax"])

    bal_gstr2a_6cut = bal_gstr2a_5cut[~mask_1a]
    bal_itr_6cut = bal_itr_5cut[~mask_1b]

    
    #NOw, after all matching, we are further analyzing the et Unmatched Cases


    print(f"Analyzing the Unmatched cases of ITR.... Please wait..!")

    #First, we check whether the PAN we are searching is present in GSTr2A at all or not. If not present then
    #we identify it separately. These have absolute no chaces of matching


    pan_itr=list(set(list(bal_itr_6cut["PAN_Number"].values)))
    pan_2a=bal_gstr2a_6cut["PAN_Number"].values

    maskpan=bal_itr_6cut["PAN_Number"].isin(pan_2a)

    bal_itr_6cut1=bal_itr_6cut[maskpan]

    unmatched_itr1=bal_itr_6cut[~maskpan]

    unmatched_itr1["Remarks_Effcorp"]="PAN/GST not available in GSTR2A"

    ws["E25"].value = len(unmatched_itr1["Remarks_Effcorp"])
    ws["F25"].value = sum(unmatched_itr1["Total_Tax"])


    #Second, we will see the CheckSUm Digit of the GST Number. Whether the last charater which is acheck sum is matching or not
    #this is also very crucials, as If GSTIN is invalid, there is no point of matching



    bal_itr_6cut1["GSTN Status"]=bal_itr_6cut1["Vendor_GST_REG"].apply(lambda x:gstchecksum(x))

    mask1=bal_itr_6cut1["GSTN Status"].values=="Check Sum MATCH"


    bal_itr_6cut2=bal_itr_6cut1[mask1]

    unmatched_itr2=bal_itr_6cut1[~mask1]


    unmatched_itr2["Remarks_Effcorp"]="GST Number Check Sum Incorrect"



    ws["E26"].value = len(unmatched_itr2["Remarks_Effcorp"])
    ws["F26"].value = sum(unmatched_itr2["Total_Tax"])


    #Third, we will be checking the Invoice Number check
    #if Invoice Number exceeds 16 digits , then we will be marking these seprately as no chaces of matching

    bal_itr_6cut2["Invoice No Check"]=bal_itr_6cut2["Invoice_Number"].apply(lambda x:gstinvcheck(x))

    mask2=bal_itr_6cut2["Invoice No Check"].values=="Invoice Number Valid"


    bal_itr_6cut3=bal_itr_6cut2[mask2]
    unmatched_itr3=bal_itr_6cut2[~mask2]

    unmatched_itr3["Remarks_Effcorp"]="Invoice No length exceed 16 digit"

    

    ws["E27"].value = len(unmatched_itr3["Remarks_Effcorp"])
    ws["F27"].value = sum(unmatched_itr3["Total_Tax"])


    #hre we will ttry to do the fuzzy matching of the Invoice Number with the GSTR2A

    print("Trying to do some Fuzzy matches in GSTR2A and ITR. Please wait....!!")


    from difflib import SequenceMatcher , get_close_matches

    cant_match=[]
    matches_itr=[]
    matches_gstr2ai=[]

    df=list(set(list(bal_itr_6cut3["PAN_Number"].values)))

    for i in df:
    #     Here the i variable is storing the PAN number each time the loop runs
        itr_balinv=bal_itr_6cut3[bal_itr_6cut3["PAN_Number"].values==i]["Invoice_Number"].values
        
        # print(f"This is ITR Invoice of {i}")
        # print(itr_balinv)
        
        gstr2a_balinv=(bal_gstr2a_6cut[bal_gstr2a_6cut["PAN_Number"].values==i]["Inv_CN_DN_Number_Finall"].values).tolist()
        # print(f"This is GSTR2A Invoice of {i}")
        # print(gstr2a_balinv)
        

        
        zipped=zip(itr_balinv,gstr2a_balinv)
        
        
        if len(gstr2a_balinv)==0:
            cant_match.append(itr_balinv)
            
        
            
        else:
               
            for inv in itr_balinv:
                
                matches_gstr2a=get_close_matches(inv,gstr2a_balinv,n=1,cutoff=0.90)
                
                if len(matches_gstr2a)==1:
                    
                    # print(f"this is inv{inv}")
                    matches_itr.append(inv)
                    matches_gstr2ai.append(matches_gstr2a[0])
                    
                    try:
                        gstr2a_balinv.remove(matches_gstr2a[0])
                    except:
                        continue
                else:
                    
                    continue
            
          
            
            cant_match.append(list(set(itr_balinv)-set(matches_itr)))
        

    mask1a=bal_itr_6cut3["Invoice_Number"].isin(matches_itr)

    mask1b=bal_gstr2a_6cut["Inv_CN_DN_Number_Finall"].isin(matches_gstr2ai)



    prob_itr_match=bal_itr_6cut3[mask1a]

    prob_gstr2a_match=bal_gstr2a_6cut[mask1b]

    prob_gstr2a_match["Matching Category"] = "Probable Match- Fuzzy Logic"
    prob_itr_match["Matching Category"] = "Probable Match- Fuzzy Logic"



    bal_itr_6cut4=bal_itr_6cut3[~mask1a]

    bal_gstr2a_7cut=bal_gstr2a_6cut[~mask1b]

    bal_itr_6cut4["Remarks_Effcorp"]="These Cases are Not Matching"


    ws["C18"].value = len(prob_gstr2a_match["Inv_CN_DN_Number_Finall"])
    ws["D18"].value = sum(prob_gstr2a_match["Total_Tax"])
    ws["E18"].value = len(prob_itr_match["Invoice_Number"])
    ws["F18"].value = sum(prob_itr_match["Total_Tax"])


    print(f"Matchig is done...Creating the 2 Files for you. Summary.xlsx & Working.xlsx")



    #now, we will be merging all these Unmatched cases of itr and final balance cut ITR

    bal_itr_7cut=pd.concat([unmatched_itr1,unmatched_itr2,unmatched_itr3,bal_itr_6cut4])


    gstr2a.to_excel(writer, sheet_name='Orignal GSTR2A', index=False)

    itr.to_excel(writer, sheet_name='Original ITR', index=False)

    all_matched_2a = pd.concat([matched_gstr2a_3way, matched_gstr2a_2way1, matched_gstr2a_2way2,matched_gstr2a_3way2, matched_gstr2a_2way3,matched_gstr2a_2way4,prob_gstr2a_match], ignore_index=True)

    all_matched_itr = pd.concat([matched_itr_3way, matched_itr_2way1, matched_itr_2way2,matched_itr_3way2,matched_itr_2way3, matched_itr_2way4,prob_itr_match], ignore_index=True)

    all_matched_2a.to_excel(writer, sheet_name='Matched_GSTR2A', index=False)

    all_matched_itr.to_excel(writer, sheet_name='Matched_ITR', index=False)

    bal_gstr2a_7cut.to_excel(writer, sheet_name='Unmatched_GSTR2A', index=False)

    bal_itr_7cut.to_excel(writer, sheet_name='Unmatched_ITR', index=False)


    ws["C28"].value = len(bal_gstr2a_7cut["GST_INVDATE_2_WAY"])
    ws["D28"].value = sum(bal_gstr2a_7cut["Total_Tax"])
    ws["E28"].value = len(bal_itr_6cut4["Remarks_Effcorp"])
    ws["F28"].value = sum(bal_itr_6cut4["Total_Tax"])



    writer.save()

    print("Success! ")



    wb.save(fullpath2)
    writer.save()

    wb.close()
    writer.close()


    print(f'Matching has been done and saved in below path \n {fullpath2}\n ')

    return (writer)






def flatten_dict(dic):
    """
    This is  avery useful function to flatten any dictionary which consists of a nested list of a nested dictionary

    This function taken only 1 parameter ,

    The parameter must be a dictionary

    This function has a depndent function i.e expand_list

    Both these functions always to be used simultaneously as they are compimentary to each other


    """
        
    import pandas as pd
    import json
    import warnings
    from openpyxl import load_workbook
    import os

    warnings.filterwarnings('ignore')

    df2 = pd.DataFrame()

    key_list = list(dic.keys())
    flat_dict = dict()

    for i in key_list:
        dict_whole = {i: dic[i]}
        dict_value = dic[i]

        if isinstance(dict_value, dict):
            flat_dict.update(dict_value)

        elif isinstance(dict_value, list):

            if len(dict_value) == 1:
                a = dict_value[0]
                b = flatten_dict(a)
                flat_dict.update(b)
            elif len(dict_value) == 0:
                pass

            else:
                dicdf = expand_list(dict_value)
                flat_dict.update(dicdf)

        else:
            flat_dict.update(dict_whole)

    data_list = list(flat_dict.items())

    df = pd.DataFrame(data_list)
    df1 = df.T
    df1.columns = df1.loc[0]
    df1 = df1.drop(0)
    df1 = df1.reset_index(drop=True)
    return (flat_dict)


def expand_list(list_dic):
    """
    This is  avery useful function to flatten any list which consists of a nested list or even a nested dictionary

    This function taken only 1 parameter ,

    The parameter must be of type list

    This function has a depndent function i.e flatten_dict

    Both these functions always to be used simultaneously as they are compimentary to each other

    """
    
    import pandas as pd
    import json
    import warnings
    from openpyxl import load_workbook
    import os

    warnings.filterwarnings('ignore')


    df2 = pd.DataFrame()

    if len(list_dic) == 1:
        a = list_dic[0]
        b = flatten_dict(a)
        conv_dict = b
    else:

        for i in list_dic:
            if isinstance(i, dict):

                flat_dictl = flatten_dict(i)

                try:
                    df1 = pd.DataFrame(flat_dictl)
                except:
                    df = pd.DataFrame(list(flat_dictl.items()))
                    df1 = df.T
                    df1.columns = df1.loc[0]
                    df1 = df1.drop(0)
                    df1 = df1.reset_index(drop=True)

                df2 = df2.append(df1)

            elif isinstance(i, list):
                a = expand_list(i)
                df2 = a

            else:
                dict_whole = {i: list_dic[i]}
                df = pd.DataFrame(list(dict_whole.items()))
                df1 = df.T
                df1.columns = df1.loc[0]
                df1 = df1.drop(0)
                df1 = df1.reset_index(drop=True)

                df2 = df2.append(df1)

        conv_dict = df2.to_dict(orient="list")

    return (conv_dict)



def gstr1_to_excel(filepath):
    """
    This is a very easy to use funcion to extract the json data of GSTR1 into an excel file.

    This function takes only one argument i.e a completepath to the json file upto extension

    Simply pass the complete path and run.

    Invoice wise data will be populated in the Excel sheet

    """

    
    import pandas as pd
    import json
    import warnings
    from openpyxl import load_workbook
    import os

    warnings.filterwarnings('ignore')


    print(f'The Json GSTR-1 file path selected is {filepath}')
    print("We are analyzing the sheets available")


    folder=os.path.dirname(filepath)
    
    fullpath1=folder+"\\"+"Converted_GSTR-1 Table Wise"+filepath.split("\\")[-1].split(".")[0]+".xlsx"

    
    pth = os.path.dirname(str(filepath))
    
    # fullpath1 = pth + "/" + "GSTR-1 Table Wise.xlsx"
    
    writer = pd.ExcelWriter(fullpath1, engine='xlsxwriter', options={'strings_to_formulas': True})

    fullpath1a=folder+"\\"+"Converted_GSTR-1 Summary"+filepath.split("\\")[-1].split(".")[0]+".xlsx"


    # fullpath1a = pth + "/" + "Summary.xlsx"
    writer1 = pd.ExcelWriter(fullpath1a, engine='xlsxwriter', options={'strings_to_formulas': True})

    df1 = pd.DataFrame()
    df1.to_excel(writer1, sheet_name="Summary_GSTR1", index=False)

    writer1.save()

    fullpath2 = fullpath1a.replace("/", "\\")  # this is a very useful command for defining the correct filepath

    wb = load_workbook(fullpath2)
    ws = wb["Summary_GSTR1"]

    ws["A1"].value = "AUTOMATION SOLUTIONS BY EFFICIENT CORPORATES-[TM]"
    ws["A4"].value = "Summary of the GSTR-1 File Compiled"

    ws["A6"].value = "GSTIN of the Seller"
    ws["A7"].value = "GSTR-1 filing period"
    ws["A8"].value = "GSTR-1 filing Type"
    ws["A9"].value = "GT"

    ws["A10"].value = "CUR_GT"
    ws["A11"].value = "FILING DATE"

    ws["A15"].value = "SUMMARY OF THE DIFFERENT TABLES IN THE GSTR-1 FILE"

    ws.merge_cells("A15:I15")

    ws["A17"].value = "GSTR-1 Tables"
    ws["B17"].value = "Count"
    ws["C17"].value = "Invoice Amount"
    ws["D17"].value = "Taxable Value"
    ws["E17"].value = "IGST"
    ws["F17"].value = "CGST"

    ws["G17"].value = "SGST"
    ws["H17"].value = "Cess"
    ws["I17"].value = "Total Tax Amount"

    ws["A18"].value = "BUSINESS-2- BUSINESS (B2B)"
    ws["A19"].value = "BUSINESS-2- CONSUMER-SMALL (B2CS)"

    ws["A20"].value = "BUSINESS-2- CONSUMER-LARGE (B2CL)"
    ws["A21"].value = "EXPORT (EXP)"
    ws["A22"].value = "CREDIT NOTE / DEBIT NOTE (CDNR)"

    ws["A24"].value = "HSN SUMMARY"

    with open(filepath) as json_file:
        data = json.load(json_file)

    dic_keys = data.keys

    for i in dic_keys():

        if i == "gstin":
            print("We are getting the Meta Data for you...Please wait...!")
            gst = data[i]
            ws["B6"].value = gst

        elif i == "fp":
            fp = data[i]
            ws["B7"].value = fp

        elif i == "filing_typ":
            fil = data[i]
            if fil == "M":
                ws["B8"].value = "Monthly"
            elif fil == "Q":
                ws["B8"].value = "Quarterly"
            else:
                ws["B8"].value = fil

        elif i == "gt":
            gt = data[i]
            ws["B9"].value = gt

        elif i == "cur_gt":
            cur_gt = data[i]
            ws["B10"].value = cur_gt

        elif i == "b2b":

            print("Fetching the B2B data, Please wait for some time...!!")
            b2b_data = data[i]
            dic_b2b = expand_list(b2b_data)
            df_b2b = pd.DataFrame(dic_b2b)
            df_b2b["GSTR-Table"] = "B2B"
            df_b2b["Json File Name"] = filepath
            df_b2b.to_excel(writer, sheet_name='B2B_DATA', index=False)

        elif i == "b2cl":

            print("Fetching the B2CL data, Please wait for some time...!!")
            b2cl_data = data[i]
            dic_b2cl = expand_list(b2cl_data)
            df_b2cl = pd.DataFrame(dic_b2cl)
            df_b2cl["GSTR-Table"] = "B2C-L"
            df_b2cl["Json File Name"] = filepath
            df_b2cl.to_excel(writer, sheet_name='B2CL_DATA', index=False)

        elif i == "cdnr":

            print("Fetching the CDNR data, Please wait for some time...!!")
            cdnr_data = data[i]
            dic_cdnr = expand_list(cdnr_data)
            df_cdnr = pd.DataFrame(dic_cdnr)
            df_cdnr["GSTR-Table"] = "CDNR"
            df_cdnr["Json File Name"] = filepath
            df_cdnr.to_excel(writer, sheet_name='CDNR_DATA', index=False)


        elif i == "exp":

            print("Fetching the Export data, Please wait for some time...!!")
            exp_data = data[i]
            dic_exp = expand_list(exp_data)
            df_exp = pd.DataFrame(dic_exp)
            df_exp["GSTR-Table"] = "EXPORT"
            df_exp["Json File Name"] = filepath
            df_exp.to_excel(writer, sheet_name='EXPORT_DATA', index=False)

        elif i == "b2cs":

            print("Fetching the B2CS data, Please wait for some time...!!")
            b2cs_data = data[i]
            dic_b2cs = expand_list(b2cs_data)
            df_b2cs = pd.DataFrame(dic_b2cs)
            df_b2cs["GSTR-Table"] = "B2C-S"
            df_b2cs["Json File Name"] = filepath
            df_b2cs.to_excel(writer, sheet_name='B2CS_DATA', index=False)

        elif i == "hsn":

            print("Getting the HSN Summary For you...!!")
            hsn_data = data[i]
            dic_hsn = flatten_dict(hsn_data)
            df_hsn = pd.DataFrame(dic_hsn)
            df_hsn.to_excel(writer, sheet_name='HSN_DATA', index=False)

        elif i == "nil":
            nil_data = data[i]
            dic_nil = flatten_dict(nil_data)
            df_nil = pd.DataFrame(dic_nil)
            df_nil.to_excel(writer, sheet_name='NIL_NONGST_DATA', index=False)

        elif i == "doc_issue":
            print("Getting the Document Series Summary For you...!!")
            doc_data = data[i]
            dic_doc = flatten_dict(doc_data)
            df_doc = pd.DataFrame(dic_doc)
            df_doc.to_excel(writer, sheet_name='DOC_SERIES_DATA', index=False)

        elif i == "fil_dt":
            fildt = data["fil_dt"]
            ws["B11"].value = fildt

        else:
            add_case = data[i]
            if isinstance(add_case, list):
                dic_add_case = expand_list(add_case)
                df_add_case = pd.DataFrame(dic_add_case)
                df_add_case["GSTR-Table"] = i
                df_add_case.to_excel(writer, sheet_name=i, index=False)
            elif isinstance(add_case, dict):
                dic_add_case = flatten_dict(add_case)
                df_add_case = pd.DataFrame(dic_add_case)
                df_add_case["GSTR-Table"] = i
                df_add_case.to_excel(writer, sheet_name=i, index=False)
            else:
                pass

    wb.save(fullpath2)
    writer.save()
    #
    # print("Consolidating All Major Tables in Single Sheet for you..!!")
    # df_comb = pd.concat([df_b2b, df_b2cl, df_cdnr, df_exp, df_b2cs])
    # df_comb.to_excel(writer, sheet_name="All_Combined_Case", index=False)

    wb.save(fullpath2)
    # writer.save()

    try:
        ws["B18"].value = len(df_b2b["ctin"])
        ws["C18"].value = df_b2b["val"].sum()
        ws["D18"].value = df_b2b["txval"].sum()
        ws["E18"].value = df_b2b["iamt"].sum()
        ws["F18"].value = df_b2b["camt"].sum()
        ws["G18"].value = df_b2b["samt"].sum()
        ws["H18"].value = df_b2b["csamt"].sum()
        ws["I18"].value = df_b2b["iamt"].sum() + df_b2b["camt"].sum() + df_b2b["samt"].sum() + df_b2b["csamt"].sum()
    except:
        pass

    
    try:

        ws["B19"].value = len(df_b2cs["rt"])
        #     ws["C19"].value = sum(df_b2cs["val"])
        ws["D19"].value = df_b2cs["txval"].sum()
        ws["E19"].value = df_b2cs["iamt"].sum()
        ws["F19"].value = df_b2cs["camt"].sum()
        ws["G19"].value = df_b2cs["samt"].sum()
        # ws["H19"].value = df_b2cs["csamt"].sum()
        ws["I19"].value = df_b2cs["iamt"].sum() + df_b2cs["camt"].sum() + df_b2cs["samt"].sum()
    except:
        pass



    try:
        ws["B20"].value = len(df_b2cl["val"])
        ws["C20"].value = df_b2cl["val"].sum()
        ws["D20"].value = df_b2cl["txval"].sum()
        ws["E20"].value = df_b2cl["iamt"].sum()
        #     ws["F20"].value = df_b2cl["camt"].sum()
        #     ws["G20"].value = df_b2cl["samt"].sum()
        ws["H20"].value = df_b2cl["csamt"].sum()
        ws["I20"].value = df_b2cl["iamt"].sum()

    except:
        pass


    try:

        ws["B21"].value = len(df_exp["flag"])
        ws["C21"].value = df_exp["val"].sum()
        ws["D21"].value = df_exp["txval"].sum()
        ws["E21"].value = df_exp["iamt"].sum()
        #     ws["F21"].value = df_exp["camt"].sum()
        #     ws["G21"].value = df_exp["samt"].sum()
        ws["H21"].value = df_exp["csamt"].sum()
        ws["I21"].value = df_exp["iamt"].sum()
    except:
        pass


    try:

        ws["B22"].value = len(df_cdnr["flag"])
        ws["C22"].value = df_cdnr["val"].sum()
        ws["D22"].value = df_cdnr["txval"].sum()
        ws["E22"].value = df_cdnr["iamt"].sum()
        #     ws["F22"].value = df_cdnr["camt"].sum()
        #     ws["G22"].value = df_cdnr["samt"].sum()
        ws["H22"].value = df_cdnr["csamt"].sum()
        ws["I22"].value = df_cdnr["iamt"].sum()

    except:
        pass

    try:
        ws["B24"].value = len(df_hsn["flag"])
        #     ws["C24"].value = df_hsn["val"].sum()
        ws["D24"].value = df_hsn["txval"].sum()
        ws["E24"].value = df_hsn["iamt"].sum()
        ws["F24"].value = df_hsn["camt"].sum()
        ws["G24"].value = df_hsn["samt"].sum()
        ws["H24"].value = df_hsn["csamt"].sum()
        ws["I24"].value = df_hsn["iamt"].sum() + df_hsn["camt"].sum() + df_hsn["samt"].sum() + df_hsn["csamt"].sum()

    except:
        pass

    writer.save()

    print("All Data have been extracted Successfully! ")

    wb.save(fullpath2)
    writer.save()

    wb.close()
    writer.close()

    

    print("We have created two Excel files for you..!! 1) Summary.xlsx and 2) GSTR-1 Table Wise.xlsx")

    print(f'The Excel Files are Extracted and kept in the below path \n {fullpath2}\n{fullpath1}\n\n ')

    return(writer)



def merge_gstr2b_excel(folder):


    """

    This function is for merging all the GSTR2B excel file as downloaded from the GST Portal.

    This function takes only one parameter i.e the folder pth in which these excel files are stored.

    Do not keep any other excel file. ALso, do no keep the GSTR2B summary file in this folder.


    """

    import pandas as pd
    import glob
    import os
    from UliPlot.XLSX import auto_adjust_xlsx_column_width
    

    filenames = glob.glob(folder + "/*.xlsx")
    
    df_master=pd.DataFrame()

  

    df_b2b=pd.DataFrame()
    
    
          
    path=os.path.join(folder + "\All_Combined_GSTR2B.xlsx")

    writer=pd.ExcelWriter(path,engine='xlsxwriter',engine_kwargs={'options': {'strings_to_numbers': True}})
  

    for file in filenames:
        
        
        file_name=file.split("\\")[-1]
        print(f"Working on B2B data...for file {file_name}")
        
        try:

            df=pd.read_excel(file,sheet_name ="B2B")
            df =df.rename({'Goods and Services Tax  - GSTR-2B':'GSTIN_of_Supplier',
                           'Unnamed: 1':'Trade_Name_of_Supplier',
                           'Unnamed: 2':'Final_Invoice_CNDN_No',
                           'Unnamed: 3':'Final_Inv_CNDN_Type',
                           'Unnamed: 4':'Final_Invoice_CNDN_Date',
                           'Unnamed: 5':'Invoice_CNDN_Value',
                           'Unnamed: 6':'Place_Of_Supply',
                           'Unnamed: 7':'Supply_Attract_Reverse_Charge',
                           'Unnamed: 8':'Tax_Rate',
                           'Unnamed: 9':'Taxable_Value',
                           'Unnamed: 10':'IGST_Amount',
                           'Unnamed: 11':'CGST_Amount',
                           'Unnamed: 12':'SGST_Amount',
                           'Unnamed: 13':'Cess_Amount',
                           'Unnamed: 14':'Supplier_Filing_Period',
                           'Unnamed: 15':'Supplier_Filing_Date',
                           'Unnamed: 16':'ITC_Available',
                           'Unnamed: 17':'Reason',
                           'Unnamed: 18':'Applicable_Percent_TaxRate',
                           'Unnamed: 19':'Source_Type',
                           'Unnamed: 20':'IRN',
                           'Unnamed: 21':'IRN_Generate_Date'}, axis=1)
            
            df =df.drop([0,1,2,3,4])
            df['GSTR2B_Table']="B2B"
            df['File_Name']=file
            df_b2b=df_b2b.append(df)
        except:
            pass
        
        
        
    
    
    df_b2ba=pd.DataFrame()
    

    for file in filenames:
        
        
        file_name=file.split("\\")[-1]
        print(f"Working on B2BA data...for file {file_name}")
        
        try:
            

            df=pd.read_excel(file,sheet_name ="B2BA")
            df =df.rename({ 'Goods and Services Tax  - GSTR-2B':'Initial_Inv_CNDN_Num',
                            'Unnamed: 1':'Initial_Inv_CNDN_Date',
                            'Unnamed: 2':'GSTIN_of_Supplier',
                            'Unnamed: 3':'Trade_Name_of_Supplier',
                            'Unnamed: 4':'Final_Invoice_CNDN_No',
                            'Unnamed: 5':'Final_Inv_CNDN_Type',
                            'Unnamed: 6':'Final_Invoice_CNDN_Date',
                            'Unnamed: 7':'Invoice_CNDN_Value',
                            'Unnamed: 8':'Place_Of_Supply',
                            'Unnamed: 9':'Supply_Attract_Reverse_Charge',
                            'Unnamed: 10':'Tax_Rate',
                            'Unnamed: 11':'Taxable_Value',
                            'Unnamed: 12':'IGST_Amount',
                            'Unnamed: 13':'CGST_Amount',
                            'Unnamed: 14':'SGST_Amount',
                            'Unnamed: 15':'Cess_Amount',
                            'Unnamed: 16':'Supplier_Filing_Period',
                            'Unnamed: 17':'Supplier_Filing_Date',
                            'Unnamed: 18':'ITC_Available',
                            'Unnamed: 19':'Reason',
                            'Unnamed: 20':'Applicable_Percent_TaxRate'}, axis=1)
            df =df.drop([0,1,2,3,4,5])
            df['GSTR2B_Table']="B2BA"
            df['File_name']=file
            df_b2ba=df_b2ba.append(df)
        except:
            pass
        
    
    
    
    df_b2bcd=pd.DataFrame()
    

    for file in filenames:
        
        
        file_name=file.split("\\")[-1]
        print(f"Working on CDNR data...for file {file_name}")
        
        try:
            

            df=pd.read_excel(file,sheet_name ="B2B-CDNR")
            df =df.rename({ 'Goods and Services Tax  - GSTR-2B':'GSTIN_of_Supplier',
                            'Unnamed: 1':'Trade_Name_of_Supplier',
                            'Unnamed: 2':'Final_Invoice_CNDN_No',
                            'Unnamed: 3':'Final_Inv_CNDN_Type',
                            'Unnamed: 4':'Note_Supply_Type',
                            'Unnamed: 5':'Final_Invoice_CNDN_Date',
                            'Unnamed: 6':'Invoice_CNDN_Value',
                            'Unnamed: 7':'Place_Of_Supply',
                            'Unnamed: 8':'Supply_Attract_Reverse_Charge',
                            'Unnamed: 9':'Tax_Rate',
                            'Unnamed: 10':'Taxable_Value',
                            'Unnamed: 11':'IGST_Amount',
                            'Unnamed: 12':'CGST_Amount',
                            'Unnamed: 13':'SGST_Amount',
                            'Unnamed: 14':'Cess_Amount',
                            'Unnamed: 15':'Supplier_Filing_Period',
                            'Unnamed: 16':'Supplier_Filing_Date',
                            'Unnamed: 17':'ITC_Available',
                            'Unnamed: 18':'Reason',
                            'Unnamed: 19':'Applicable_Percent_TaxRate',
                            'Unnamed: 20':'Source_Type',
                            'Unnamed: 21':'IRN',
                            'Unnamed: 22':'IRN_Generate_Date'}, axis=1)
            df =df.drop([0,1,2,3,4])
            df['GSTR2B_Table']="B2B-CDNR"
            df['File_name']=file
            df_b2bcd=df_b2bcd.append(df)
        except:
            pass
        
    
     
    
    df_b2bcdnra=pd.DataFrame()
    

    for file in filenames:
        
        
        
        file_name=file.split("\\")[-1]
        print(f"Working on CDNRA data...for file {file_name}")
        
        try:
            

            df=pd.read_excel(file,sheet_name ="B2B-CDNRA")
            df =df.rename({ 'Goods and Services Tax  - GSTR-2B':'Initial_Inv_CNDN_Type',
                            'Unnamed: 1':'Initial_Inv_CNDN_No',
                            'Unnamed: 2':'Initial_Inv_CNDN_Date',
                            'Unnamed: 3':'GSTIN_of_Supplier',
                            'Unnamed: 4':'Trade_Name_of_Supplier',
                            'Unnamed: 5':'Final_Invoice_CNDN_No',
                            'Unnamed: 6':'Final_Inv_CNDN_Type',
                            'Unnamed: 7':'Note_Supply_Type',
                            'Unnamed: 8':'Final_Invoice_CNDN_Date',
                            'Unnamed: 9':'Invoice_CNDN_Value',
                            'Unnamed: 10':'Place_Of_Supply',
                            'Unnamed: 11':'Supply_Attract_Reverse_Charge',
                            'Unnamed: 12':'Tax_Rate',
                            'Unnamed: 13':'Taxable_Value',
                            'Unnamed: 14':'IGST_Amount',
                            'Unnamed: 15':'CGST_Amount',
                            'Unnamed: 16':'SGST_Amount',
                            'Unnamed: 17':'Cess_Amount',
                            'Unnamed: 18':'Supplier_Filing_Period',
                            'Unnamed: 19':'Supplier_Filing_Date',
                            'Unnamed: 20':'ITC_Available',
                            'Unnamed: 21':'Reason',
                            'Unnamed: 22':'Applicable_Percent_TaxRate'}, axis=1)
            df =df.drop([0,1,2,3,4,5])
            df['GSTR2B_Table']="B2B-CDNRA"
            df['File_name']=file
            df_b2bcdnra=df_b2bcdnra.append(df)
        
        except:
            pass
        

    
    df_isd=pd.DataFrame()
    

    for file in filenames:
        
        
        
        file_name=file.split("\\")[-1]
        print(f"Working on ISD data...for file {file_name}")
        
        try:
            

            df=pd.read_excel(file,sheet_name ="ISD")
            df =df.rename({ 'Goods and Services Tax  - GSTR-2B':'GSTIN_of_Supplier',
                            'Unnamed: 1':'Trade_Name_of_Supplier',
                            'Unnamed: 2':'Final_Inv_CNDN_Type',
                            'Unnamed: 3':'Final_Invoice_CNDN_No',
                            'Unnamed: 4':'Final_Invoice_CNDN_Date',
                            'Unnamed: 5':'Initial_Inv_CNDN_No',
                            'Unnamed: 6':'Initial_Inv_CNDN_Date',
                            'Unnamed: 7':'IGST_Amount',
                            'Unnamed: 8':'CGST_Amount',
                            'Unnamed: 9':'SGST_Amount',
                            'Unnamed: 10':'Cess_Amount',
                            'Unnamed: 11':'Supplier_Filing_Period',
                            'Unnamed: 12':'Supplier_Filing_Date',
                            'Unnamed: 13':'ITC_Available'}, axis=1)
            df =df.drop([0,1,2,3,4])
            df['GSTR2B_Table']="ISD"
            df['File_name']=file
            df_isd=df_isd.append(df)
        except:
            pass
        
    
    

        

    df_impg=pd.DataFrame()
    

    for file in filenames:
        
        
        file_name=file.split("\\")[-1]
        print(f"Working on IMPORT data...for file {file_name}")
        
        try:
            

            df=pd.read_excel(file,sheet_name ="IMPG")
            df =df.rename({ 'Goods and Services Tax  - GSTR-2B':'IceGate_Ref_Date',
                            'Unnamed: 1':'Port_Code',
                            'Unnamed: 2':'Bill_Of_Entry_No',
                            'Unnamed: 3':'Record_Date',
                            'Unnamed: 4':'Taxable_Value',
                            'Unnamed: 5':'IGST_Amount',
                            'Unnamed: 6':'Cess_Amount',
                            'Unnamed: 7':'Amended_Y_N'}, axis=1)
            df =df.drop([0,1,2,3,4])
            df['File_name']=file
            df['GSTR2B_Table']="IMPG"
            df_impg=df_impg.append(df)
        except:
            pass
        
    
    
    
    print("Combining all the files...")
    

    
    
        
    df_master=pd.concat([df_b2b,df_b2ba,df_b2bcd,df_b2bcdnra,df_isd,df_impg])
    
#     df_master=df_master[['GSTIN_of_Supplier','Trade_Name_of_Supplier','Final_Invoice_CNDN_No','Final_Inv_CNDN_Type',
#                          'Final_Invoice_CNDN_Date','Revised Invoice number','Revised Invoice Date','Original Note Number',
#                          'Original Note Date','Note type','Original Note type','Note Supply type','Note Value (₹)',
#                          'Revised Note number','Revised Note date','Invoice Value(₹)','Place of supply',
#                          'Supply Attract Reverse Charge','Rate(%)','Taxable Value (₹)','Integrated Tax(₹)',
#                          'Central Tax(₹)','State/UT Tax(₹)','Cess(₹)','GSTR-1/IFF/GSTR-5 Period',
#                          'GSTR-1/IFF/GSTR-5 Filing Date','ITC Availability','Reason','Applicable % of Tax Rate',
#                          'Source','IRN','IRN Date','GSTIN of ISD','ISD Document type','ISD Document number',
#                          'ISD Document date','Original invoice date','ISD GSTR-6 Period','ISD GSTR-6 Filing Date',
#                          'Eligibility of ITC','Icegate Reference Date','Port Code','Number','Date',
#                          'Icegate Taxable Value','Amended (Yes)','Sheet name','File_name']]
    
#     df_master=df_master[["GSTIN_of_Supplier","Trade_Name_of_Supplier","GSTR2B_Table","Supply_Attract_Reverse_Charge",
#                          "Final_Invoice_CNDN_No","Final_Inv_CNDN_Type","Final_Invoice_CNDN_Date","Invoice_CNDN_Value",
#                          "Place_Of_Supply","Tax_Rate",
#                          "Taxable_Value","IGST_Amount","CGST_Amount","SGST_Amount","Cess_Amount",
#                          "Supplier_Filing_Period","Supplier_Filing_Date","ITC_Available","Reason",
#                          "Applicable_Percent_TaxRate"]]
    

        
    
    df_master.to_excel(writer,sheet_name="All_Combined",index=False)
    auto_adjust_xlsx_column_width(df_master, writer, sheet_name="All_Combined", margin=10)
    
    
    
    df_b2ba.to_excel(writer,sheet_name="B2BA",index=False)
    auto_adjust_xlsx_column_width(df_b2ba, writer, sheet_name="B2BA", margin=10)
    
    
    
    df_b2b.to_excel(writer,sheet_name="B2B",index=False)
    auto_adjust_xlsx_column_width(df_b2b, writer, sheet_name="B2B", margin=10)
 
    df_impg.to_excel(writer,sheet_name="IMPG",index=False)
    auto_adjust_xlsx_column_width(df_b2bcd, writer, sheet_name="IMPG", margin=10)    
    
        
    df_isd.to_excel(writer,sheet_name="ISD",index=False)
    auto_adjust_xlsx_column_width(df_b2bcd, writer, sheet_name="ISD", margin=10)
    


    df_b2bcdnra.to_excel(writer,sheet_name="B2B-CDNRA",index=False)
    auto_adjust_xlsx_column_width(df_b2bcd, writer, sheet_name="B2B-CDNRA", margin=10)
    
    df_b2bcd.to_excel(writer,sheet_name="B2B-CDNR",index=False)
    auto_adjust_xlsx_column_width(df_b2bcd, writer, sheet_name="B2B-CDNR", margin=10)
    
    writer.save()
    
    print(f"All excel files of GSTR2B has been Combined. Combined files stored in {path}")




def rename_2b_columns(dataframe):

    """

    This is a support function for the gstr2B to Excel conversion


    """
    
    dataframe.rename(columns={"dt":"Final_Invoice_CNDN_Date",
                "val":"Invoice_CNDN_Value",
                "rev":"Supply_Attract_Reverse_Charge",
                "itcavl":"ITC_Available",
                "diffprcnt":"Applicable_Percent_TaxRate",
                "pos":"Place_Of_Supply",
                "typ":"Final_Inv_CNDN_Type",
                "inum":"Final_Invoice_CNDN_No",
                "rsn":"Reason",
                "sgst":"SGST_Amount",
                "rt":"Tax_Rate",
                "num":"Check_num",
                "txval":"Taxable_Value",
                "cgst":"CGST_Amount",
                "cess":"Cess_Amount",
                "trdnm":"Trade_Name_of_Supplier",
                "supfildt":"Supplier_Filing_Date",
                "supprd":"Supplier_Filing_Period",
                "ctin":"GSTIN_of_Supplier",
                "igst":"IGST_Amount",
                "irn":"IRN",
                "irngendate":"IRN_Generate_Date",
                "srctyp":"Source_Type",
                "GSTR2B-Table":"GSTR2B-Table",
                "rtnprd":"GSTR2B_Period",
                "gstin":"Recipient_GSTIN",
                "Json File Name":"JSON_Source_File",
                "File_Name":"Source_Excel_File",
                "oinum":"Initial_Inv_CNDN_No",
                "oidt":"Initial_Inv_CNDN_Date",
                "ntnum":"Final_Invoice_CNDN_No",
                "suptyp":"Note_Supply_Type",
                "ontdt":"Initial_Inv_CNDN_Date",
                "onttyp":"Initial_Inv_CNDN_Type",
                "ontnum":"Initial_Inv_CNDN_No",
                "docnum":"Final_Invoice_CNDN_No",
                "itcelg":"ITC_Available",
                "doctyp":"Final_Inv_CNDN_Type",
                "docdt":"Final_Invoice_CNDN_Date",
                "oinvnum":"Initial_Inv_CNDN_No",
                "oinvdt":"Initial_Inv_CNDN_Date",
                "boedt":"Bill_Of_Entry_Date",
                "isamd":"Amended_Y_N",
                "recdt":"Record_Date",
                "refdt":"IceGate_Ref_Date",
                "boenum":"Bill_Of_Entry_No",
                "portcode":"Port_Code"},inplace=True)
    


def gstr2b_to_excel(filepath):
    """
    This is a very easy to use funcion to extract the json data of GSTR2b into an excel file.
    This function takes only one argument i.e a completepath to the json file upto extension
    Simply pass the complete path and run.
    Table wise data will be populated in the Excel sheet
    """

    import pandas as pd
    import json
    import warnings
    from openpyxl import load_workbook
    import os

    warnings.filterwarnings('ignore')

    original_name=filepath.split("\\")[-1]
    folder = os.path.dirname(filepath)
    newfile = folder + "\\"+ "Converted_Excel_"  + filepath.split("\\")[-1].split(".")[0] + ".xlsx"

    print(f"the file {original_name} has been selected... Working on it..!")
    
    
    writer = pd.ExcelWriter(newfile, engine='xlsxwriter', options={'strings_to_formulas': True})

#     writer=pd.ExcelWriter(newfile,engine='openpyxl')


    df_impg=pd.DataFrame()
    df_isd=pd.DataFrame()
    df_cdnr=pd.DataFrame()
    df_cdnra=pd.DataFrame()
    df_b2b=pd.DataFrame()
    df_b2ba=pd.DataFrame()

    with open(filepath) as json_file:
        data = json.load(json_file)

        main_data=data["data"]['docdata']

        return_period=data["data"]["rtnprd"]
        rec_gstin=data["data"]["gstin"]

        # print(abc)



        for i in main_data.keys():
#             print(i)

            if i =="b2b":

                print(f"Fetching the {i} data, Please wait for some time...!!")
                b2b_data = main_data[i]
                dic_b2b = expand_list(b2b_data)

                try:

                    df_b2b = pd.DataFrame(dic_b2b)
                except ValueError:
                    df_b2b = pd.DataFrame(dic_b2b,index=[0])


                df_b2b["GSTR2B_Table"] = i
                df_b2b["rtnprd"]=return_period
                df_b2b["gstin"]=rec_gstin
                df_b2b["Json_File_Name"]=filepath
                
                rename_2b_columns(df_b2b)
                

                
                df_b2b.to_excel(writer, sheet_name=str(i+'_data'), index=False)
                print(f"{i} Data converted to Excel....!!")

#                 writer.save()

            elif i=="b2ba":

                print(f"Fetching the {i} data, Please wait for some time...!!")
                b2ba_data = main_data[i]
                dic_b2ba = expand_list(b2ba_data)

                try:
                    df_b2ba = pd.DataFrame(dic_b2ba)
                except ValueError:
                    df_b2ba = pd.DataFrame(dic_b2ba, index=[0])


                df_b2ba["GSTR2B_Table"] = i
                df_b2ba["rtnprd"]=return_period
                df_b2ba["gstin"]=rec_gstin
                df_b2ba["Json File Name"]=filepath
                
                
                rename_2b_columns(df_b2ba)
                
                
                df_b2ba.to_excel(writer, sheet_name=str(i+'_data'), index=False)
                print(f"{i} Data converted to Excel....!!")

#                 writer.save()


            elif i=="cdnr":
                print(f"Fetching the {i} data, Please wait for some time...!!")
                cdnr_data = main_data[i]
                dic_cdnr = expand_list(cdnr_data)

                try:
                    df_cdnr = pd.DataFrame(dic_cdnr)
                except ValueError:
                    df_cdnr = pd.DataFrame(dic_cdnr, index=[0])

                df_cdnr["GSTR2B_Table"] = i
                df_cdnr["rtnprd"] = return_period
                df_cdnr["gstin"] = rec_gstin
                df_cdnr["Json File Name"] = filepath
                
                
                rename_2b_columns(df_cdnr)
                
                
                df_cdnr.to_excel(writer, sheet_name=str(i + '_data'), index=False)
                print(f"{i} Data converted to Excel....!!")

#                 writer.save()


            elif i=="cdnra":
                print(f"Fetching the {i} data, Please wait for some time...!!")
                cdnra_data = main_data[i]
                dic_cdnra = expand_list(cdnra_data)

                try:
                    df_cdnra = pd.DataFrame(dic_cdnra)
                except ValueError:
                    df_cdnra = pd.DataFrame(dic_cdnra, index=[0])


                df_cdnra["GSTR2B_Table"] = i
                df_cdnra["rtnprd"] = return_period
                df_cdnra["gstin"] = rec_gstin
                df_cdnra["Json File Name"] = filepath
                
                rename_2b_columns(df_cdnra)
                
                df_cdnra.to_excel(writer, sheet_name=str(i + '_data'), index=False)
                print(f"{i} Data converted to Excel....!!")

#                 writer.save()


            elif i=="isd":

                print(f"Fetching the {i} data, Please wait for some time...!!")
                isd_data = main_data[i]
                dic_isd = expand_list(isd_data)


                try:
                    df_isd = pd.DataFrame(dic_isd)
                except ValueError:
                    df_isd = pd.DataFrame(dic_isd, index=[0])


                df_isd["GSTR2B_Table"] = i
                df_isd["rtnprd"] = return_period
                df_isd["gstin"] = rec_gstin
                df_isd["Json File Name"] = filepath
                
                rename_2b_columns(df_isd)
                
                df_isd.to_excel(writer, sheet_name=str(i + '_data'), index=False)
                print(f"{i} Data converted to Excel....!!")

#                 writer.save()

            elif i=="impg":
                print(f"Fetching the {i} data, Please wait for some time...!!")
                impg_data = main_data[i]
                dic_impg = expand_list(impg_data)

                try:
                    df_impg = pd.DataFrame(dic_impg)
                except ValueError:
                    df_impg = pd.DataFrame(dic_impg, index=[0])

                df_impg["GSTR2B_Table"] = i
                df_impg["rtnprd"] = return_period
                df_impg["gstin"] = rec_gstin
                df_impg["Json File Name"] = filepath
                
                rename_2b_columns(df_impg)
                
                df_impg.to_excel(writer, sheet_name=str(i + '_data'), index=False)
                print(f"{i} Data converted to Excel....!!")

#                 writer.save()

            else:
                pass

        print(f"All Sheets Have been created separateky in same Excel File named {newfile}")


    combined_2b=pd.concat([df_b2b,df_b2ba,df_cdnr,df_cdnra,df_isd,df_impg])

    print("Combining all sheets into one single Excel Sheet , Named 'effcorp_all_combined'")


    combined_2b.to_excel(writer,sheet_name="effcorp_all_combined",index=False)

    print(f"Please Wait...Saving the single final file as {newfile}")

    writer.save()
    writer.close()

    return (writer)



